#!/usr/bin/env python3
"""
    File name: addsMeanDepth.py
    Author: Mariane Cote
    Date created: 02/13/2019
    Python Version: 3.6

    Searches the mean depth for each lake of the lake's list in the SHYPE data and
        adds a column with those values into the CSV file.

"""
import xlrd
from openpyxl import load_workbook

csvf = r'2017SwedenList.csv'
with open(csvf, 'rU') as f:
    lines = f.readlines()
    nlines = len(lines)
    # nlines = 3
    ii = range(1, nlines)


table = load_workbook(r'2017SwedenList.xlsx')
sheettab = table.active

sheettab.cell(row=1, column=10).value = 'depth.mean'
for i in ii:
    lake_id, subid, name, ebh, area, depth, longitude, latitude, volume\
        = lines[i].strip().split(',')

    filepath = r"C:\Users\Marianne\Documents\Fish_niche1\lake_subset_SHMI_data\SHYPEdata\%s.xls" % subid

    wb = xlrd.open_workbook(filepath)
    sheet_names = wb.sheet_names()
    sheet = wb.sheet_by_name(sheet_names[2])

    try:
        cell = sheet.cell_value(5, 1)
        print(cell)
        sheettab.cell(row=i+1, column=10).value = cell
    except:
        print("%s is missing" % subid)
    # save workbook
table.save(r'2017SwedenListnew.xlsx')
