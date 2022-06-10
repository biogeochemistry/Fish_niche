#!/usr/bin/env python

""" Script for MyLake
script specific to the visualisation of the data.
"""

__author__ = "Marianne Cote"

import matplotlib.pyplot as plt
from joblib import Parallel, delayed
from scipy.stats import linregress
import seaborn as sns
import pandas as pd
import os
import time
import math
from datetime import timedelta,date,datetime
import skill_metrics as sm
from matplotlib.ticker import FormatStrFormatter, MultipleLocator
from matplotlib.dates import YearLocator, MonthLocator, DateFormatter, DayLocator, WeekdayLocator, MONDAY
import statsmodels.api as smodels
from statsmodels.sandbox.regression.predstd import wls_prediction_std
import scipy.stats as stats
import pandas as pd
import numpy as np
from math import sqrt, floor, log10, log
from sklearn.metrics import r2_score, mean_squared_error
import xlsxwriter
# import Main_fish_niche as mainfn
# import load_workbook
from openpyxl import load_workbook
#from lake_information import LakeInfo
import statistics
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
from itertools import product
import csv
from matplotlib.offsetbox import AnchoredText
from mpl_toolkits.axes_grid1 import make_axes_locatable
# every monday
mondays = WeekdayLocator(MONDAY)
weekFormatter = DateFormatter('%b')

models = {1: ('ICHEC-EC-EARTH', 'r1i1p1_KNMI-RACMO22E_v1_day'),
          2: ('ICHEC-EC-EARTH', 'r3i1p1_DMI-HIRHAM5_v1_day'),
          3: ('MPI-M-MPI-ESM-LR', 'r1i1p1_CLMcom-CCLM4-8-17_v1_day'),
          4: ('MOHC-HadGEM2-ES', 'r1i1p1_SMHI-RCA4_v1_day'),
          5: ('IPSL-IPSL-CM5A-MR', 'r1i1p1_IPSL-INERIS-WRF331F_v1_day'),
          6: ('CNRM-CERFACS-CNRM-CM5', 'r1i1p1_CLMcom-CCLM4-8-17_v1_day')}

scenarios = {1: ('historical', 1971, 'historical', 1976),
             2: ('historical', 2001, 'rcp45', 2006),
             3: ('rcp45', 2031, 'rcp45', 2036),
             4: ('rcp45', 2061, 'rcp45', 2066),
             5: ('rcp45', 2091, 'rcp45', 2096),
             6: ('rcp85', 2031, 'rcp85', 2036),
             7: ('rcp85', 2061, 'rcp85', 2066),
             8: ('rcp85', 2091, 'rcp85', 2096)}



def sums_of_squares(obs_list, sims_list):
    """
    Finds the sums of squares for all temperatures listed in the comparison file.
    :param obs_list: A list of observed temperatures.
    :param sims_list: A list of simulated temperatures.
    :return: The result of the sums of square as a float.
    """
    sums = 0
    for x in range(len(obs_list)):
        if obs_list[x] == 'None':
            continue
        sums += (float(obs_list[x]) - float(sims_list[x])) ** 2

    return sums

def root_mean_square(obs_list, sims_list):
    """
    Finds the root_mean_square for the temperatures listed in the comparison file.
    :param obs_list: A list of observed temperatures.
    :param sims_list: A list of simulated temperatures.
    :return: The result of the root mean square as a float.
    """

    d = pd.DataFrame(list(zip(obs_list, sims_list)),
                     columns=['obs', 'sim'])

    d["obs"] = d["obs"].astype(float)
    d["sim"] = d["sim"].astype(float)

    try:
        results = mean_squared_error(d["obs"], d["sim"])
        results = sqrt(results)
        resultsnormalise = sqrt(mean_squared_error(d["obs"], d["sim"])) / (max(d["obs"]) - min(d["obs"]))
    except:
        results = np.nan
        resultsnormalise = np.nan
    return results, resultsnormalise

def r_squared(obs_list, sims_list):
    """
    Find the R squared for the simulations compared to the expected observations
    :param obs_list: A list of observed temperatures.
    :param sims_list: A list of simulated temperatures.
    :return: results of R squared, as a float
    """

    x = []
    y = []

    for i in obs_list:
        try:

            x.append(float(i))
            y.append(float(sims_list[obs_list.index(i)]))

        except ValueError:
            continue
        except IndexError:
            break
    try:
        rsquare = r2_score(x, y)
    except:
        rsquare = np.nan
    return rsquare

def standard_deviation(obs_list):
    """
    Find the standard deviation of the observations
    :param obs_list: Type list. The list of observed temperatures
    :return: The standard deviation of obs_list
    """
    observations = []
    for obs in obs_list:
        try:
            observations.append(float(obs))
        except ValueError:
            continue

    return statistics.stdev(observations)

def rmse_by_sd(obs_list, rmse):
    """
    Divides RMSE of the simulations by the SD of the observations
    :param obs_list: A list of observed temperatures.
    :param rmse: Float
    :return: A float, RMSE / SD
    """
    try:
        results = rmse / standard_deviation(obs_list)
    except ZeroDivisionError:
        results = "Error_Zero_Division"
    return results

def round_decimals_down(number:float, decimals:int=2):
    """
    Returns a value rounded down to a specific number of decimal places.
    """
    if not isinstance(decimals, int):
        raise TypeError("decimal places must be an integer")
    elif decimals < 0:
        raise ValueError("decimal places has to be 0 or more")
    elif decimals == 0:
        return math.floor(number)

    factor = 10 ** decimals
    return math.floor(number * factor) / factor

def round_decimals_up(number:float, decimals:int=2):
    """
    Returns a value rounded up to a specific number of decimal places.
    """
    if not isinstance(decimals, int):
        raise TypeError("decimal places must be an integer")
    elif decimals < 0:
        raise ValueError("decimal places has to be 0 or more")
    elif decimals == 0:
        return math.ceil(number)

    factor = 10 ** decimals
    return math.ceil(number * factor) / factor

def Area_at_depth():
    lakes = pd.read_csv("2017SwedenList_only_validation_12lakes.csv", encoding='ISO-8859-1')
    lakes_small = lakes[lakes["volume"]< 1.0e7]['lake_id'].tolist()
    lakes_medium = lakes[lakes["volume"]>= 1.0e7 ]
    lakes_medium = lakes_medium[lakes_medium['volume'] <= 5.0e9]['lake_id'].tolist()
    lakes_large = lakes[lakes["volume"] >  5.0e9]['lake_id'].tolist()

    T_list = [2, 4, 6, 8, 10, 12, 13, 14, 15]
    light_list = [0.5, 1, 2, 4, 8, 16, 24, 32, 48]
    AreaDays=[]
    Areaday=[]
    NTG=[]
    for taille in ['small','medium','large']:
        for x in T_list:
            AreaDaysT = []
            AreadayT = []
            NTGT = []
            for y in light_list:
                if y == 0.5:
                    y = '0.50'
                df = pd.read_csv(os.path.join("F:\output", "fish_niche_Area_Light%s_T%s_2001-2010.csv"%(y,x)))
                if taille == 'small':
                    dataset = df[df['LakeId'].isin(lakes_small)].mean()
                elif taille== 'medium':
                    dataset = df[df['LakeId'].isin(lakes_medium)].mean()
                else:
                    dataset = df[df['LakeId'].isin(lakes_large)].mean()
                AreaDaysT.append(dataset['AreaDays'])
                AreadayT.append(dataset['Areaday'])
                NTGT.append(dataset['NTGdays'])
            AreaDays.append(AreaDaysT)
            Areaday.append(AreadayT)
            NTG.append(NTGT)

        x = np.array(T_list)
        y = np.array(light_list)

        #x = np.linspace(0, 5, 50)
        #y = np.linspace(0, 5, 40)
        X, Y = np.meshgrid(x, y)
        #Z = f(X, Y)

        ZAreaday = np.array(Areaday)

        ZAreadays = np.array(AreaDays)
        ZNTG = np.array(NTG)


        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)

        plt.imshow(ZAreaday, extent=[2, 15, 0.5, 48], origin='lower', cmap='RdGy', interpolation='nearest', aspect='auto')
        plt.colorbar()

        plt.savefig("Areaday_%s.png" % taille)
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        plt.imshow(ZAreadays, extent=[2, 15, 0.5, 48], origin='lower', cmap='ocean', interpolation='nearest', aspect='auto')
        plt.colorbar()
        plt.savefig("Areadays_%s.png" % taille)
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        plt.imshow(ZNTG, extent=[2, 15, 0.5, 48], origin='lower', cmap='bone', interpolation='nearest', aspect='auto')
        plt.colorbar()
        plt.axis(aspect='image')
        plt.savefig("NTG_%s.png" % taille)
        plt.close()

def timeline_plot(modeleddata: list, modeleddates: list, observeddata: list = None, observeddates: list = None, ax=None,ylimit=[-0.5,30],
                  line_kwargs: dict = {},
                  sct_kwargs: dict = {},linestyle = "-"):
    """
    plot timeline with modeled data (line) and observed data (dot)

    :param modeleddata: modeled data for each day included in years presenting observed data (at same depth)
    :param modeleddates: all days included in the period modeled
    :param observeddata: Observed data to be plot
    :param observeddates: Dates Where observed data have been measured
    :param ax: Axe where you want to plot. If None, look for last axe used in current figure
    :param line_kwargs: Others arguments given to the lineplot (measures' plot)
    :param sct_kwargs: Others arguments given to the scatterplot (Observations' plot)
    :return: None
    """
    if ax is None:
        ax = plt.gca()
    if modeleddates is None or modeleddata is None:
        raise TypeError

    if observeddata is not None:
        sns.scatterplot(x=observeddates, y=observeddata, ax=ax, **sct_kwargs)

    sns.lineplot(x=modeleddates, y=modeleddata, ax=ax, **line_kwargs)

    ax.set_xlabel("Dates")
    ax.set_xlim(min(modeleddates), max(modeleddates))
    ax.set_ylim(ylimit[0],ylimit[1])

def target_diagram_background_custom(crmsd, bias, max_value_overall, min_value_overall, ax=None, sct_args={},
                              cicle_max_args = {}, cicle_min_args = {}, cicle_1_args = {}, other_circle_args = {},
                              space_between_circles = 0.5, max_number_cicle = 3, add_circles = True,color=['k','k','k'],label=['','',''],size = [10,10,10]):
    if ax is None:
        ax = plt.gca()

    if min_value_overall > max_value_overall:
        max_value = min_value_overall
        min_value = max_value_overall
    else:
        max_value = max_value_overall
        min_value = min_value_overall

    if add_circles:
        ax.add_patch(plt.Circle((0, 0), max_value, **cicle_max_args))
        ax.add_patch(plt.Circle((0, 0), 1, **cicle_1_args))
        ax.add_patch(plt.Circle((0, 0), min_value, **cicle_min_args))

        if min_value < 0.2:
            new_circle = math.floor(min_value)
        else:
            new_circle = math.ceil(min_value)
        for i in range(max_number_cicle-1):
            if (max_value-new_circle) > space_between_circles:
                if (new_circle+space_between_circles) != 1:
                    ax.add_patch(plt.Circle((0, 0), new_circle, **other_circle_args))
                    new_circle += space_between_circles
            else:
                break

    for i in range(len(crmsd)):
        sct_args['color'] = color[i]
        sct_args['s'] = size[i]
        sct_args['label'] = label[i]
        plt.scatter(y=bias[i], x=crmsd[i], **sct_args)

    if max_value < 1:
        limit = 1 + space_between_circles
    else:
        limit = max_value + space_between_circles

    plt.ylim(-1 * limit, limit)
    plt.xlim(-1 * limit, limit)
    sns.despine()
    ax.spines['left'].set_position('center')
    ax.spines['bottom'].set_position('center')
    plt.xlabel('Normalized\n RMSE', horizontalalignment='right', x=1)
    # #ax.set_xlabel('Normalized cRMSE', loc='right')
    plt.ylabel('Normalized Bias', verticalalignment="top", y=1)
    # ax.set_ylabel('Normalized Bias')
    ax.xaxis.set_minor_locator(MultipleLocator(5))
    ax.yaxis.set_minor_locator(MultipleLocator(2))
    ax.set_aspect(1.0)

def target_diagram_background(crmsd, bias, max_value_overall, min_value_overall, ax=None, sct_args={},
                              cicle_max_args = {}, cicle_min_args = {}, cicle_1_args = {}, other_circle_args = {},
                              space_between_circles = 0.2,max_number_cicle = 3, add_circles = True,color=['k','k','k'],label=['','',''],size = [10,10,10],edgecolor= ['k','k','k']):
    if ax is None:
        ax = plt.gca()

    if min_value_overall > max_value_overall:
        max_value = min_value_overall
        min_value = max_value_overall
    else:
        max_value = max_value_overall
        min_value = min_value_overall

    if add_circles:
        if max_value > 1.5:
            limit = 4
            ax.add_patch(plt.Circle((0, 0), 3.5, **cicle_max_args))
            # ax.add_patch(plt.Circle((0, 0), 3, **other_circle_args))
            # ax.add_patch(plt.Circle((0, 0), 2, **other_circle_args))
            ax.add_patch(plt.Circle((0, 0), 2.5, **other_circle_args))
            ax.add_patch(plt.Circle((0, 0), 1.5, **other_circle_args))
            ax.add_patch(plt.Circle((0, 0), 1, **cicle_1_args))
            ax.add_patch(plt.Circle((0, 0), 0.5, **cicle_min_args))
        else:
            limit = 2
            ax.add_patch(plt.Circle((0, 0), 1.5, **cicle_max_args))
            ax.add_patch(plt.Circle((0, 0), 1, **cicle_1_args))
            ax.add_patch(plt.Circle((0, 0), 0.5, **cicle_min_args))
    else:
        if max_value < 1:
            limit = 1 + space_between_circles
        else:
            limit = max_value + space_between_circles

        # if min_value < 0.2:
        #     new_circle = math.floor(min_value)
        # else:
        #     new_circle = math.ceil(min_value)
        # for i in range(max_number_cicle-1):
        #     if (max_value-new_circle) > space_between_circles:
        #         if (new_circle+space_between_circles) != 1:
        #             ax.add_patch(plt.Circle((0, 0), new_circle, **other_circle_args))
        #             new_circle += space_between_circles
        #     else:
        #         break

    for i in range(len(crmsd)):
        sct_args['facecolor'] = color[i]
        sct_args['edgecolor'] = edgecolor[i]
        sct_args['s'] = size[i]
        sct_args['label'] = label[i]
        plt.scatter(y=bias[i], x=crmsd[i], **sct_args)




    plt.ylim(-1 * limit, limit)
    plt.xlim(-1 * limit, limit)
    sns.despine()
    ax.spines['left'].set_position('center')
    ax.spines['bottom'].set_position('center')
    plt.xlabel('Normalized\n RMSE', horizontalalignment='right', x=0.2)
    # #ax.set_xlabel('Normalized cRMSE', loc='right')
    plt.ylabel('Normalized Bias\n', verticalalignment="top", y=0.9,labelpad=10)
    # ax.set_ylabel('Normalized Bias')
    ax.xaxis.set_minor_locator(MultipleLocator(5))
    ax.yaxis.set_minor_locator(MultipleLocator(5))
    ax.set_aspect(1.0)

def line_plot(lineStart=None,lineEnd=None, ax=None,linearg={},font_family="Times New Roman"):
    if ax is None:
        ax = plt.gca()
    if lineStart is None:
        lineStart = 0
        lineEnd = 20

    plt.plot([lineStart, lineEnd], [lineStart, lineEnd], **linearg)
    plt.xlim(lineStart, lineEnd)
    plt.ylim(lineStart, lineEnd)
    plt.rcParams.update({"font.family": font_family})

def linear_regression_plot(x,y,ax=None,linearregressionarg={}, confidentintervalarg = {}, predictionintervalarg = {},font_family = "Times New Roman"):
    if ax is None:
        ax = plt.gca()


    x = smodels.add_constant(x)  # constant intercept term
    # Model: y ~ x + c
    model = smodels.OLS(y, x)
    fitted = model.fit()
    x_pred = np.linspace(x.min(), x.max(), 50)
    x_pred2 = smodels.add_constant(x_pred)
    y_pred = fitted.predict(x_pred2)

    ax.plot(x_pred, y_pred, **linearregressionarg)

    print(fitted.params)  # the estimated parameters for the regression line
    print(fitted.summary())  # summary statistics for the regression

    y_hat = fitted.predict(x)  # x is an array from line 12 above
    y_err = y - y_hat
    mean_x = x.T[1].mean()
    n = len(x)
    dof = n - fitted.df_model - 1

    t = stats.t.ppf(1 - 0.025, df=dof)
    s_err = np.sum(np.power(y_err, 2))
    conf = t * np.sqrt((s_err / (n - 2)) * (1.0 / n + (
            np.power((x_pred - mean_x), 2) / ((np.sum(np.power(x_pred, 2))) - n * (np.power(mean_x, 2))))))
    upper = y_pred + abs(conf)
    lower = y_pred - abs(conf)

    #Confidence Interval
    ax.fill_between(x_pred, lower, upper, **confidentintervalarg)

    #Prediction Interval
    sdev, lower, upper = wls_prediction_std(fitted, exog=x_pred2, alpha=0.025)
    ax.fill_between(x_pred, lower, upper, **predictionintervalarg)
    plt.rcParams.update({"font.family": font_family})

def error_bar_plot(x,y,xerr,yerr,ax=None, errorbararg = {},markerwidth = 1,SIZE=10,font_family="Times New Roman"):
    if ax is None:
        ax = plt.gca()

    (_, caps, _) = plt.errorbar(x, y, xerr=xerr, yerr=yerr, **errorbararg)
    for cap in caps:
        cap.set_markeredgewidth(markerwidth)

    plt.rcParams['font.size'] = '%s' % SIZE
    plt.rc('xtick', labelsize=SIZE)  # fontsize of the tick labels
    plt.rc('ytick', labelsize=SIZE)
    plt.rcParams.update({'font.size': SIZE})
    plt.rcParams.update({"font.family": font_family})

def base_plot_comparison(x,y, lineStart=None,lineEnd= None, ax=None,linecolor = "k",bigger=30,font_family="Arial"):
    if ax is None:
        ax = plt.gca()
    plt.rcParams.update({"font.family": font_family})
    plt.rcParams['font.family'] = 'serif'
    plt.rcParams['font.serif'] = [font_family] + plt.rcParams['font.serif']
    plt.rcParams['mathtext.default'] = 'regular'
    linearg = { 'color': linecolor, 'label': "y= x", 'linewidth': 1,'linestyle':'--'}
    slope, intercept, r_value, p_value, std_err = linregress(x, y)
    rmse, nrmse = root_mean_square(x, y)
    linearregressionarg = { 'color': 'k', 'linewidth': 1,
                        "label": "linear regression (y = %0.3f x + %0.3f) \n R\u00b2 : %0.3f RMSE: %0.3f" % (slope, intercept, r_value, rmse)}
    confidenceintervalarg = {'color': '#888888', 'alpha': 0.4, 'label': "Confidence interval"}
    predictionintervalarg = {'color': '#888888', 'alpha': 0.1, 'label': "Prediction interval"}
    line_plot(lineStart=lineStart,lineEnd=lineEnd,ax=ax, linearg=linearg,font_family=font_family)
    ax.text(lineStart+0.5,lineEnd-0.5, "$ R\u00b2 : %0.3f \ RMSE: %0.3f $" % (r_value, rmse), horizontalalignment='left',verticalalignment='top',fontsize=bigger)
    linear_regression_plot(x, y, ax, linearregressionarg, confidenceintervalarg, predictionintervalarg,font_family=font_family)
    plt.rcParams.update({"font.family":font_family})

def base_contourplot(X,Y,z_variable_data, z_variablelist,lake,number_of_plot_by_row,vmin,vmax,ax = None, colorlinecontour = "white",individual=False, ntg=False):

    if type(z_variable_data) is dict:
        list_z = range(0, len(z_variablelist))
        for z in list_z:
            if ntg:
                Z=[]
                for list in z_variable_data["%s" % z_variablelist[z]]:
                    Z.append([365 if i >= 365 else i for i in list])
            else:
                Z = z_variable_data["%s" % z_variablelist[z]]
            if ax is None:
                axs = plt.gca()
            else:
                try:
                    axis_x = int(np.floor(z / number_of_plot_by_row))
                    axis_y = int(z- (number_of_plot_by_row * axis_x))
                    axs = ax[axis_x][axis_y]
                except:
                    axs = ax

            try:
                cp1 = axs.contourf(X, Y, Z, vmin=0, vmax=365)

                cp = axs.contour(X, Y, Z, colors=colorlinecontour,linewidths=1)
                axs.tick_params(axis='both', pad=-0.1)
                plt.setp(axs.yaxis.get_ticklines(), 'markersize', 3)
                plt.setp(axs.xaxis.get_ticklines(), 'markersize', 3)

                # plt.setp(axs.yaxis.get_ticklines(), 'markeredgewidth', 1)

                # plt.subplot_tool()
                # plt.tight_layout()

                if ntg:
                    plt.clabel(cp, inline=True,fmt='%.0f')
                else:
                    plt.clabel(cp, inline=True, fmt='%.2e')
                axs.title.set_text("DO threshold: %s mg/l" % z_variablelist[z])
            except:
                print("issue: possible number for subplot higher than number for subplot present")


        if len(list_z) < 16 and len(list_z) != 1:
            axs[3][3].set_visible(False)

        return cp1
    else:
        if individual:
            if ax is None:
                axs = plt.gca()
            else:
                axs = ax
            cm = axs.contourf(X, Y, z_variable_data, vmin=vmin, vmax=vmax)
            cp = axs.contour(X, Y, z_variable_data, colors=colorlinecontour)
            plt.colorbar(cm)  # Add a colorbar to a plot
            plt.clabel(cp, inline=True, fmt='%.2e')
            ax.title.set_text("DO threshold: %s mg/l" % z_variablelist[0])
        return None

def create_dictionnary_of_z_value_by_lake(data,z_list,list_of_parameter_measured,lake_id_list,new_data):
    vmin,vmax = {},{}
    for z_value in [str(x) for x in z_list]:
        for column in [str(x) for x in list_of_parameter_measured]:
            for lake_id in [str(x) for x in lake_id_list]:
                if not column in new_data:
                    new_data[column]= {lake_id: {z_value:data[z_value][column][lake_id]}}
                else:
                    if not lake_id in new_data[column]:
                        new_data[column][lake_id] = {z_value: data[z_value][column][lake_id]}
                    else:
                        if not z_value in new_data[column][lake_id]:
                            new_data[column][lake_id][z_value]= data[z_value][column][lake_id]
                        else:
                            new_data[column][lake_id][z_value].append(data[z_value][column][lake_id])
    for column in [str(x) for x in list_of_parameter_measured]:
        for lake_id in [str(x) for x in lake_id_list]:
            vminlakemethod = []
            vmaxlakemethod = []
            if not column in vmin:
                for z_value in [str(x) for x in z_list]:
                    vminlakemethod.append(min([min(x) for x in new_data[column][lake_id][z_value]]))
                    vmaxlakemethod.append(max([max(x) for x in new_data[column][lake_id][z_value]]))

                vmin[column] = {lake_id: min(vminlakemethod)}
                vmax[column] = {lake_id: max(vmaxlakemethod)}
            else:
                for z_value in [str(x) for x in z_list]:
                    vminlakemethod.append(min([min(x) for x in new_data[column][lake_id][z_value]]))
                    vmaxlakemethod.append(max([max(x) for x in new_data[column][lake_id][z_value]]))

                vmin[column][lake_id]= min(vminlakemethod)
                vmax[column][lake_id]= max(vmaxlakemethod)


    return new_data,vmin,vmax

def get_file_by_value_of_xyz(x_value,y_value,z_value,dict_raw_data_by_lake,variables_list_in_order,directory,lake_id_list):
    #file format:
    #Need to be change if other file naming is use

    variables_in_order_for_file = ['Light', 'Temperature', 'Oxygen']
    general_name_file = "fish_niche_Area"
    simulated_period = "2001-2010"

    order_for_file = variables_in_order_for_file
    order_for_file[variables_in_order_for_file.index(variables_list_in_order[0])] = x_value
    order_for_file[variables_in_order_for_file.index(variables_list_in_order[1])] = y_value
    order_for_file[variables_in_order_for_file.index(variables_list_in_order[2])] = z_value
    variables_in_order_for_file = ['Light', 'Temperature', 'Oxygen']

    filename = "%s_%s%s_%s%s_%s%s_%s.csv"%(general_name_file,variables_in_order_for_file[0],order_for_file[0],variables_in_order_for_file[1],order_for_file[1],variables_in_order_for_file[2],order_for_file[2],simulated_period)


    # Open and extract data from file
    with open('%s/%s' % (directory, filename), newline='') as f:
        reader = csv.reader(f)
        data = list(reader)

    list_of_parameter_measured = data[0][1:]
    list_validate = ["%s"% x for x in lake_id_list]
    for lake in range(0,len(data)):

        if "%s" % data[lake][0] in  list_validate:
            if not "%s" % data[lake][0] in dict_raw_data_by_lake:
                try:
                    dict_raw_data_by_lake["%s" % data[lake][0]] = [[float(data[lake][x]) for x in range(1,len(data[lake]))]]
                except ValueError:
                    dict_raw_data_by_lake["%s" % data[lake][0]] = [[data[lake][x] for x in range(1,len(data[lake]))]]
            else:
                try:
                    dict_raw_data_by_lake["%s" % data[lake][0]].append([float(data[lake][x]) for x in range(1,len(data[lake]))])
                except ValueError:
                    dict_raw_data_by_lake["%s" % data[lake][0]].append([data[lake][x] for x in range(1,len(data[lake]))])

    return dict_raw_data_by_lake, list_of_parameter_measured

class Graphics:
    """
    Main class that regroup all function to plot modeled data
    """

    def __init__(self, outputfolder,width=3.25, height= 3.25,font_family = "Times New Roman",size = 12):
        self.output_folder = outputfolder
        self.time = time.strftime("%Y%m%d-%H%M%S")
        # self.time = 4
        self.SMALL_SIZE = size-2
        self.MEDIUM_SIZE = size-1
        self.BIGGER_SIZE = size
        self.font_family = font_family
        plt.style.context('seaborn-paper')
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams["font.family"] = font_family
        plt.rcParams.update({"font.family": font_family})
        self.width, self.height = width, height

    def comparison_obs_sims_plot(self, variable_analized, calibration_methods, modeldata, obsdata, depthlayers,
                                 ice_cover,lake_name, icecovertarea=True):
        """

        :type obsdata: pd.DataFrame
        :param variable_analized:
        :param calibration_methods:
        :param modeldata:
        :param obsdata:
        :param depthlayers:
        :param ice_cover:
        :param icecovertarea:
        :return:
        """

        # Figures aesthetics
        colorsSchemeByWaterLevel = {"surface": [ "#B30000","black", "#FF9785"],
                                    "deepwater": [ '#14218F',"black", "#0AEFFF"]}
        subplotAxeByWaterLevel = {"surface": 0, "deepwater": 1}
        markerStyleByVariable = {"Temperature": "s", "Oxygen Concentration": "o"}
        markerColorByWaterLevel = {"surface": colorsSchemeByWaterLevel['surface'][0],
                                   "deepwater": colorsSchemeByWaterLevel['deepwater'][0]}
        lineStyleByWaterLevel = {"GA2": "-", "GA1": "--", "SR": "-."}
        sns.set_style("ticks", {"xtick.major.size": 100, "ytick.major.size": 100})
        plt.xticks(rotation=15)
        linewidthByMethod = {"GA2": 1.5, "GA1": 2, "SR": 1}
        orderPlotPresentedByMethod = {"Observation": 100, "GA2": 10, "GA1": 50,
                                      "SR": 80}
        scatterPlotDotSize = 50
        transparenceLineplot = 0.8
        iceCovertAreaColor = "grey"
        iceCovertAreaAlpha = 0.2
        legendNames = [ "First GA Calibration","Second GA Calibration", "Stewise Regression", "Observations"]

        fig, axs = plt.subplots(2, 1, constrained_layout=True, figsize=(15, 8), gridspec_kw={'height_ratios': [1, 1]})

        for depth_level in ["surface", "deepwater"]:
            axe_level = subplotAxeByWaterLevel[depth_level]
            depthlayer = depthlayers[depth_level]

            if not obsdata.loc[obsdata['Depth'] == depthlayer]["Observations"].empty:

                # Ice Covert Area
                if icecovertarea:
                    secondaryAxe = axs[axe_level].twinx()
                    secondaryAxe.fill_between(ice_cover.iloc[:, 1].tolist(), ice_cover.iloc[:, 0].tolist(),
                                              color=iceCovertAreaColor, alpha=iceCovertAreaAlpha, zorder=-10)
                    secondaryAxe.set_ylim(0, 1)
                    secondaryAxe.set_yticklabels([])
                    secondaryAxe.yaxis.set_visible(False)

                # Plot each method and observations
                for method in calibration_methods:
                    lineplotstyle = {"linewidth": linewidthByMethod[method],
                                     "color": colorsSchemeByWaterLevel[depth_level][calibration_methods.index(method)],
                                     "zorder": orderPlotPresentedByMethod[method],
                                     "alpha": transparenceLineplot}

                    if method != "SR":
                        timeline_plot(modeleddata=modeldata["%s_Model_%s" % (method, depth_level)],
                                      modeleddates=modeldata['Dates'], ax=axs[axe_level],
                                      line_kwargs=lineplotstyle)


                        axs[axe_level].lines[calibration_methods.index(method)].set_linestyle(lineStyleByWaterLevel[method])

                    else:
                        scatterplotstyle = {'marker': markerStyleByVariable[variable_analized],
                                            's': scatterPlotDotSize, "edgecolor": 'k',
                                            "facecolors": markerColorByWaterLevel[depth_level],
                                            "linewidth": linewidthByMethod[method],
                                            "linestyle": "-",
                                            "zorder": orderPlotPresentedByMethod["Observation"]}

                        timeline_plot(modeleddata=modeldata["%s_Model_%s" % (method, depth_level)],
                                      modeleddates=modeldata['Dates'],
                                      observeddata=obsdata.loc[obsdata['Depth'] == depthlayer]["Observations"],
                                      observeddates=obsdata.loc[obsdata['Depth'] == depthlayer]["Dates"],
                                      ax=axs[axe_level], sct_kwargs=scatterplotstyle,line_kwargs=lineplotstyle)
                        axs[axe_level].lines[calibration_methods.index(method)].set_linestyle(
                            lineStyleByWaterLevel[method])

                # set y axis title to subplot in function of the variable analysed and the depth 
                if variable_analized == "Temperature":
                    unity = "°C"
                    axs[axe_level].set_ylabel("%s at %s m (%s)" % (variable_analized, depthlayer, unity))
                else:
                    unity = "mg*$L^-1$"
                    axs[axe_level].set(ylabel="%s \n at %s m (mg*$\mathregular{L^{\-1}}$)" % (variable_analized, depthlayer))
                # legend
                axs[axe_level].legend(legendNames)
                legend = axs[axe_level].legend(legendNames)
                legend.get_frame().set_facecolor('white')

        # Add title to figure
        # plt.title('timeline_lake_%s_%s'%(self.lake_name,variable_analized)))

        # Save figure
        if variable_analized == "Oxygen Concentration":
            variable_analized = "Oxygen"
        plt.savefig(
            os.path.join(self.output_folder,
                         "Comparative_timeline_%s_%s_%s" % (lake_name, variable_analized, self.time)))
        plt.close('all')


    def comparison_obs_sims_plot_V2(self, variable_analized, calibration_methods, modeldata, obsdata, depthlayers,
                                 ice_cover,lake_name, icecovertarea=True,observation=True):
        """

        :type obsdata: pd.DataFrame
        :param variable_analized:
        :param calibration_methods:
        :param modeldata:
        :param obsdata:
        :param depthlayers:
        :param ice_cover:
        :param icecovertarea:
        :return:
        """

        # Figures aesthetics
        # width, height = 6.5,3.5
        colorsSchemeByWaterLevel = {"surface": [ "black" ,"#B30000"],
                                    "deepwater": [ "black",'#14218F']}
        subplotAxeByWaterLevel = {"surface": 0, "deepwater": 1}
        markerStyleByVariable = {"Temperature": "s", "DO Concentration": "o"}
        markerColorByWaterLevel = {"surface": colorsSchemeByWaterLevel['surface'][1],
                                   "deepwater": colorsSchemeByWaterLevel['deepwater'][1]}
        lineStyleByWaterLevel = {"GA": "-",  "SR": "-."}
        # sns.set_style("ticks", {"xtick.major.size": 100, "ytick.major.size": 100})

        linewidthByMethod = {"GA": 1, "SR": 1}
        orderPlotPresentedByMethod = {"Observation": 100, "GA": 10,
                                      "SR": 80}
        scatterPlotDotSize = 30
        transparenceLineplot = 0.8
        iceCovertAreaColor = "grey"
        iceCovertAreaAlpha = 0.2
        if observation:
            legendNames = [ "GA", "SR", "Observations"]
        else:
            legendNames = ["Stepwise Regression"]


        # fig, axs = plt.subplots(2, 1, figsize=(15, 8))#, gridspec_kw={'height_ratios': [1, 1]})

        fig, axs = plt.subplots(2, 1, constrained_layout=True, figsize=(15, 8), gridspec_kw={'height_ratios': [1, 1]})
        fig = plt.gcf()
        fig.set_size_inches(self.width, self.height)
        plt.rcParams["font.family"] = self.font_family
        if variable_analized == "DO Concentration":
            limit=[-0.5,30]
        else:
            limit=[-0.5,35]
        results = pd.DataFrame(columns=[])
        stop = False
        for depth_level in ["surface", "deepwater"]:
            axe_level = subplotAxeByWaterLevel[depth_level]

            try:
                depthlayer = depthlayers[depth_level]

            except:
                stop = True

            if not stop:
                if not obsdata.loc[obsdata['Depth'] == depthlayer]["Observations"].empty:

                    # Ice Covert Area
                    if icecovertarea:
                        secondaryAxe = axs[axe_level].twinx()
                        secondaryAxe.fill_between(ice_cover.iloc[:, 1].tolist(), ice_cover.iloc[:, 0].tolist(),
                                                  color=iceCovertAreaColor, alpha=iceCovertAreaAlpha, zorder=-10)
                        secondaryAxe.set_ylim(0, 1)
                        secondaryAxe.set_yticklabels([])
                        secondaryAxe.yaxis.set_visible(False)

                    result = pd.DataFrame()
                    result["Dates"] = obsdata.loc[obsdata['Depth'] == depthlayer]["Dates"]
                    result['Observations'] = obsdata.loc[obsdata['Depth'] == depthlayer]["Observations"]
                    result.set_index('Dates', inplace = True)

                    # Plot each method and observations
                    for method in calibration_methods:
                        lineplotstyle = {"linewidth": linewidthByMethod[method],
                                         "color": colorsSchemeByWaterLevel[depth_level][calibration_methods.index(method)],
                                         "zorder": orderPlotPresentedByMethod[method],
                                         "alpha": transparenceLineplot}


                        model_data = pd.DataFrame()
                        model_data["%s_Model_%s" % (method, depth_level)] = modeldata["%s_Model_%s" % (method, depth_level)]
                        model_data['Dates'] = modeldata['Dates']
                        model_data.set_index('Dates')
                        # concact = pd.concat([result,model_data], ignore_index=True, sort=True)
                        # concact = model_data.append(result, ignore_index=True, sort=True)
                        concact = pd.concat([model_data, result], axis=1)
                        result = concact




                        if method != "SR" and observation:
                            timeline_plot(modeleddata=modeldata["%s_Model_%s" % (method, depth_level)],
                                          modeleddates=modeldata['Dates'], ax=axs[axe_level],ylimit=limit,
                                          line_kwargs=lineplotstyle)


                            axs[axe_level].lines[calibration_methods.index(method)].set_linestyle(lineStyleByWaterLevel[method])


                        else:
                            scatterplotstyle = {'marker': markerStyleByVariable[variable_analized],
                                                's': scatterPlotDotSize, "edgecolor": 'k',
                                                "facecolors": markerColorByWaterLevel[depth_level],
                                                "linewidth": linewidthByMethod[method],
                                                "linestyle": "-",
                                                "zorder": orderPlotPresentedByMethod["Observation"]}

                            timeline_plot(modeleddata=modeldata["%s_Model_%s" % (method, depth_level)],
                                          modeleddates=modeldata['Dates'],
                                          observeddata=obsdata.loc[obsdata['Depth'] == depthlayer]["Observations"],
                                          observeddates=obsdata.loc[obsdata['Depth'] == depthlayer]["Dates"],
                                          ax=axs[axe_level],ylimit=limit, sct_kwargs=scatterplotstyle,line_kwargs=lineplotstyle)
                            axs[axe_level].lines[calibration_methods.index(method)].set_linestyle(
                                lineStyleByWaterLevel[method])



                    results = pd.concat([results, result], axis=1)
                    results = results.loc[:, ~results.columns.duplicated()]
                    # results = pd.concat([results, result], ignore_index=True, sort=False)
                    # set y axis title to subplot in function of the variable analysed and the depth
                    if variable_analized == "Temperature":
                        unity = "°C"
                        axs[axe_level].set_ylabel("%s \n at %s m (°C)" % (variable_analized, depthlayer))
                    else:
                        unity = "mg*$L^-1$"
                        variable = "DO concentration"
                        # axs[axe_level].set( ylabel="%s \n at %s m (mg*$\mathregular{L^{\-1}}$)" % (variable_analized, depthlayer))
                        axs[axe_level].set_ylabel("%s \n at %s m (mg*$\mathregular{L^{-1}}$)" % (variable, depthlayer), linespacing=0.8)

                    # legend
                    legends = axs[axe_level].legend(legendNames,ncol= len(legendNames))
                    legends.get_frame().set_facecolor('#FFFFFF')
                else:

                    # Ice Covert Area
                    if icecovertarea:
                        secondaryAxe = axs[axe_level].twinx()
                        secondaryAxe.fill_between(ice_cover.iloc[:, 1].tolist(), ice_cover.iloc[:, 0].tolist(),
                                                  color=iceCovertAreaColor, alpha=iceCovertAreaAlpha, zorder=-10)
                        secondaryAxe.set_ylim(0, 1)
                        secondaryAxe.set_yticklabels([])
                        secondaryAxe.yaxis.set_visible(False)

                    result = pd.DataFrame()
                    result["Dates"] = obsdata.loc[obsdata['Depth'] == depthlayer]["Dates"]
                    result['Observations'] = obsdata.loc[obsdata['Depth'] == depthlayer]["Observations"]
                    result.set_index('Dates', inplace = True)

                    # Plot each method and observations
                    for method in calibration_methods:
                        lineplotstyle = {"linewidth": linewidthByMethod[method],
                                         "color": colorsSchemeByWaterLevel[depth_level][calibration_methods.index(method)],
                                         "zorder": orderPlotPresentedByMethod[method],
                                         "alpha": transparenceLineplot}


                        model_data = pd.DataFrame()
                        model_data["%s_Model_%s" % (method, depth_level)] = modeldata["%s_Model_%s" % (method, depth_level)]
                        model_data['Dates'] = modeldata['Dates']
                        model_data.set_index('Dates')
                        # concact = pd.concat([result,model_data], ignore_index=True, sort=True)
                        # concact = model_data.append(result, ignore_index=True, sort=True)
                        concact = pd.concat([model_data, result], axis=1)
                        result = concact




                        if method != "SR" and observation:
                            timeline_plot(modeleddata=modeldata["%s_Model_%s" % (method, depth_level)],
                                          modeleddates=modeldata['Dates'], ax=axs[axe_level],ylimit=limit,
                                          line_kwargs=lineplotstyle)


                            axs[axe_level].lines[calibration_methods.index(method)].set_linestyle(lineStyleByWaterLevel[method])

                        else:
                            scatterplotstyle = {'marker': markerStyleByVariable[variable_analized],
                                                's': scatterPlotDotSize, "edgecolor": 'k',
                                                "facecolors": markerColorByWaterLevel[depth_level],
                                                "linewidth": linewidthByMethod[method],
                                                "linestyle": "-",
                                                "zorder": orderPlotPresentedByMethod["Observation"]}

                            timeline_plot(modeleddata=modeldata["%s_Model_%s" % (method, depth_level)],
                                          modeleddates=modeldata['Dates'],
                                          observeddata=obsdata.loc[obsdata['Depth'] == depthlayer]["Observations"],
                                          observeddates=obsdata.loc[obsdata['Depth'] == depthlayer]["Dates"],
                                          ax=axs[axe_level],ylimit=limit, sct_kwargs=scatterplotstyle,line_kwargs=lineplotstyle)
                            axs[axe_level].lines[calibration_methods.index(method)].set_linestyle(
                                lineStyleByWaterLevel[method])


                    results = pd.concat([results, result], axis=1)
                    results = results.loc[:, ~results.columns.duplicated()]
                    # results = pd.concat([results, result], ignore_index=True, sort=False)
                    # set y axis title to subplot in function of the variable analysed and the depth
                    if variable_analized == "Temperature":
                        unity = "°C"
                        axs[axe_level].set_ylabel("%s \n at %s m (°C)" % (variable_analized, depthlayer))
                    else:
                        unity = "mg*$L^-1$"
                        variable = "Oxygen"
                        # axs[axe_level].set( ylabel="%s \n at %s m (mg*$\mathregular{L^{\-1}}$)" % (variable_analized, depthlayer))
                        axs[axe_level].set_ylabel("%s \n at %s m (mg*$\mathregular{L^{-1}}$)" % (variable, depthlayer), linespacing=0.8)

                    # legend
                    legends = axs[axe_level].legend(legendNames,ncol= len(legendNames))
                    legends.get_frame().set_facecolor('#FFFFFF')
            axs[axe_level].set_xticklabels([2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008, 2009, 2010], rotation=(0), va='bottom', ha='center')
            axs[axe_level].set_yticklabels([" 0"," 0", "10", "20", "30"],va= 'center', ha='left')
            if depth_level == "surface":
                axs[axe_level].xaxis.set_label_text('foo')
                axs[axe_level].xaxis.label.set_visible(False)
        plt.style.context('seaborn-paper')
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams["font.family"] = self.font_family
        plt.tight_layout(h_pad=0.5,rect=[0.05, 0.05, 0.05,0.05])

        fig = plt.gcf()
        fig.set_size_inches(self.width, self.height)
        # plt.tight_layout(pad=0.6,rect=[0.02, 0.02, 1, 1])
        # Add title to figure
        # plt.title('timeline_lake_%s_%s'%(self.lake_name,variable_analized)))

        if observation:
            # Save figure
            if variable_analized == "DO Concentration":
                variable_analized = "Oxygen1"
            plt.savefig(
                os.path.join(self.output_folder,
                             "Comparative_timeline_%s_%s_%s%s.svg" % (lake_name, variable_analized, self.time,self.font_family)))
            plt.savefig(
                os.path.join(self.output_folder,
                             "Comparative_timeline_%s_%s_%s%s.eps" % (
                             lake_name, variable_analized, self.time, self.font_family)),format='eps')
            plt.savefig(
                os.path.join(self.output_folder,
                             "Comparative_timeline_%s_%s_%s%s.png" % (lake_name, variable_analized, self.time,self.font_family)))
            plt.close('all')

        else:
            # Save figure
            if variable_analized == "DO Concentration":
                variable_analized = "Oxygen1"
            # plt.savefig(
            #     os.path.join(self.output_folder,
            #                  "Comparative_timeline_%s_%s_%s.svg" % (lake_name, variable_analized, self.time)))
            plt.savefig(
                os.path.join(self.output_folder,
                             "Comparative_timeline_%s_%s_%s%s.png" % (lake_name, variable_analized, self.time,self.font_family)))
            plt.close('all')
        plt.clf()


        return(results)

    def taylor_target_plot(self,all_data_from_model, all_data_from_observation, label_method, variable, information,
                           label_taylor):
        """

        :param all_data_from_model:
        :param all_data_from_observation:
        :param label_method:
        :param variable:
        :return:
        """
        # take list of lists
        # [[[]*3]*12]

        colorpalette = ['#7A0014', '#E7FFD6', '#4A7FBF']  # sns.color_palette('Accent_r', 3)
        colorpalette = ['grey','black','white']
        color_circle = "k"

        if variable == "Temperature":
            markers = ["s"]*12
        else:
            markers = ["o"]*12
        # markers = ["o", "v", "^", "s", "P", "*", ">", "X", "D", "<", "p", "d"]
        sizes = [200, 150, 100]
        circle_argsmin = {"color": color_circle, "fill": False, "ls": '--', 'zorder': 0}
        circle_argsmax = {"color": color_circle, "fill": False, "ls": '-.', 'zorder': 0}
        circle_args = {"color": color_circle, "fill": False, "ls": '-', 'zorder': 0}
        other = {"color": "grey", "fill": False, "ls": ':', 'zorder': 0}

        biaslist, crmsdlist, rmsdlist = [None] * (len(all_data_from_model)), [None] * (len(all_data_from_model)), [
            None] * (len(all_data_from_model))
        sdevlist, crmsdtaylist, ccoeflist = [None] * (len(all_data_from_model)), [None] * (len(all_data_from_model)), [
            None] * (len(all_data_from_model))
        fig1, ax1 = plt.subplots(figsize=(10, 10), dpi=100)
        print(len(all_data_from_model))
        for lake in range(0, len(all_data_from_model)):
            observations = all_data_from_observation[lake]
            models = all_data_from_model[lake]
            print(len(models))
            statistic_target_lake = [None] * (len(models))
            statistic_taylor_lake = [None] * (len(models) + 1)
            bias, crmsd, rmsd = [None] * (len(models)), [None] * (len(models)), [None] * (len(models))
            sdev, crmsdtay, ccoef = [None] * (len(models) + 1), [None] * (len(models) + 1), [None] * (len(models) + 1)
            for method in range(0, len(models)):
                stats_target = sm.target_statistics(models[method], observations[method], norm=True)
                stats_taylor = sm.taylor_statistics(models[method], observations[method])

                # Calculate statistics for target and taylor diagram
                bias[method] = stats_target['bias']
                crmsd[method] = stats_target['crmsd']
                rmsd[method] = stats_target['rmsd']

                if method == 0:
                    crmsdtay[method] = stats_taylor['crmsd'][0]
                    ccoef[method] = stats_taylor['ccoef'][0]
                    sdev[method] = stats_taylor['sdev'][0] / np.mean(observations[method])

                crmsdtay[method + 1] = stats_taylor['crmsd'][1]
                ccoef[method + 1] = stats_taylor['ccoef'][1]
                sdev[method + 1] = stats_taylor['sdev'][1] / np.mean(models[method])

                statistic_target_lake[method] = stats_target
                statistic_taylor_lake[method] = stats_taylor


            try:
                maxv = abs(round_decimals_up(max([max(bias, key=abs),max(crmsd, key=abs)])))
            except:
                maxv = 1
            try:
                minv = abs(round_decimals_down(min([min(bias, key=abs), min(crmsd, key=abs)])))
            except:
                minv = 0
            if not information == "all_lakes":
                sct_args = {'color': colorpalette[method], 'marker': markers[lake],
                            'label': label_method[method], 'zorder': 100, 'alpha': 0.8, 's': sizes[method], 'edgecolors': 'k'}
                target_diagram_background(crmsd=crmsd,bias=bias, max_value_overall=maxv, min_value_overall=minv, ax=ax1,max_number_cicle = 5,
                                    add_circles=True, sct_args=sct_args,cicle_max_args = circle_argsmax,
                                    cicle_min_args = circle_argsmin, cicle_1_args = circle_args, other_circle_args = other,color=colorpalette,label=label_method,size = sizes)

            else:
                sct_args = {'color': colorpalette[method], 'marker': markers[lake],
                            'label': label_method[method], 'zorder': 100, 'alpha': 0.8, 's': sizes[method],
                            'edgecolors': 'k'}
                biaslist[lake], crmsdlist[lake], rmsdlist[lake] = bias, crmsd, rmsd
                sdevlist[lake], crmsdtaylist[lake], ccoeflist[lake] = sdev, crmsdtay, ccoef

                print(len(all_data_from_model))
                if lake !=(len(all_data_from_model)-1):
                    target_diagram_background(crmsd=crmsd, bias=bias, max_value_overall=maxv,min_value_overall=minv, ax=ax1,max_number_cicle = 5, add_circles=False,sct_args=sct_args,cicle_max_args = circle_argsmax,
                                    cicle_min_args = circle_argsmin, cicle_1_args = circle_args, other_circle_args = other,color=colorpalette,label=label_method,size = sizes)
                else:
                    t1 = [max(i) for i in biaslist]
                    t2 = max(t1)
                    t3 = max([t2, max([max(i) for i in crmsdlist])])
                    try:
                        # maxv = abs(
                        # round_decimals_up(max([max([max(i) for i in biaslist]), max([max(i) for i in crmsdlist])])))
                        maxbias = [max(bias, key=abs)for bias in biaslist]
                        maxbiass = max(maxbias, key=abs)
                        maxcrmsd = [max(crmsd, key=abs) for crmsd in crmsdlist]
                        maxcrmsds = max(maxcrmsd, key=abs)
                        maxx = max([maxbiass,maxcrmsds],key=abs)
                        maxr = round_decimals_up(maxx)
                        maxv = abs(maxr)
                        # maxv = abs(round_decimals_up( max([max([max(bias, key=abs)for bias in biaslist], key=abs),max([max(crmsd, key=abs) for crmsd in crmsdlist], key=abs)], key=abs)))
                    except:
                        maxv = 1
                    try:
                        minbias = [min(bias, key=abs) for bias in biaslist]
                        minbiass = min(minbias, key=abs)
                        mincrmsd = [min(crmsd, key=abs) for crmsd in crmsdlist]
                        mincrmsds = min(mincrmsd, key=abs)
                        minx = min([minbiass, mincrmsds], key=abs)
                        minr = round_decimals_down(minx)
                        minv = abs(minr)
                        # minv = abs(
                        #     round_decimals_down(
                        #         min([min([min(i) for i in biaslist]), min([min(i) for i in crmsdlist])])))
                    except:
                        minv = 0

                    sct_args = {'color': colorpalette[method], 'marker': markers[lake],
                            'label': label_method[method], 'zorder': 100, 'alpha': 0.8, 's': sizes[method], 'edgecolors': 'k'}
                    target_diagram_background( crmsd=crmsd, bias=bias, max_value_overall=maxv, min_value_overall=minv,
                                        ax=ax1, max_number_cicle = 5,add_circles=True,sct_args=sct_args,cicle_max_args = circle_argsmax,
                                        cicle_min_args = circle_argsmin, cicle_1_args = circle_args,
                                        other_circle_args = other,color=colorpalette,label=label_method,size = sizes)
            if lake == 0:
                plt.legend()
                if information == "all_lakes":
                    biaslist[lake], crmsdlist[lake], rmsdlist[lake] = bias, crmsd, rmsd
                    sdevlist[lake], crmsdtaylist[lake], ccoeflist[lake] = sdev, crmsdtay, ccoef

        # plt.show()

        plt.savefig(
            os.path.join(self.output_folder, "Comparative_target_%s_%s_%s" % (variable, information, self.time)))
        plt.close()

    def taylor_target_plot_v2(self, all_data_from_model, all_data_from_observation, label_method, variable, information,
                           label_taylor,speciallakes=[None]):
        """

        :param all_data_from_model:
        :param all_data_from_observation:
        :param label_method:
        :param variable:
        :return:
        """
        # take list of lists
        # [[[]*3]*12]


        width, height = 3.25,3.25
        colorpalette = ['#7A0014', '#E7FFD6', '#4A7FBF']  # sns.color_palette('Accent_r', 3)
        colorpalette = [ 'black', 'white']
        special = ['cyan','red']
        color_circle = "k"

        if variable == "Temperature":
            markers = ["s"] * 12
        else:
            markers = ["o"] * 12
        # markers = ["o", "v", "^", "s", "P", "*", ">", "X", "D", "<", "p", "d"]
        sizes = [40, 30, 20]
        circle_argsmin = {"color": color_circle, "fill": False, "ls": '--', 'zorder': 0}
        circle_argsmax = {"color": color_circle, "fill": False, "ls": '-.', 'zorder': 0}
        circle_args = {"color": color_circle, "fill": False, "ls": '-', 'zorder': 0}
        other = {"color": "grey", "fill": False, "ls": ':', 'zorder': 0}
        circle_argsmin = {"color": "grey", "fill": False, "ls": ':', 'zorder': 0}
        circle_argsmax = {"color": "grey", "fill": False, "ls": ':', 'zorder': 0}

        biaslist, crmsdlist, rmsdlist = [None] * (len(all_data_from_model)), [None] * (len(all_data_from_model)), [
            None] * (len(all_data_from_model))
        sdevlist, crmsdtaylist, ccoeflist = [None] * (len(all_data_from_model)), [None] * (len(all_data_from_model)), [
            None] * (len(all_data_from_model))
        fig1, ax1 = plt.subplots(figsize=(10, 10), dpi=100)
        print(len(all_data_from_model))
        for lake in range(0, len(all_data_from_model)):
            observations = all_data_from_observation[lake]
            models = all_data_from_model[lake]
            print(len(models))
            statistic_target_lake = [None] * (len(models))
            statistic_taylor_lake = [None] * (len(models) + 1)
            bias, crmsd, rmsd = [None] * (len(models)), [None] * (len(models)), [None] * (len(models))
            sdev, crmsdtay, ccoef = [None] * (len(models) + 1), [None] * (len(models) + 1), [None] * (len(models) + 1)
            print(len(models))
            for method in range(0, len(models)):
                f = models[method]
                d = observations[method]
                stats_target = sm.target_statistics(models[method], observations[method], norm=True)
                stats_taylor = sm.taylor_statistics(models[method], observations[method])

                # Calculate statistics for target and taylor diagram
                bias[method] = stats_target['bias']
                crmsd[method] = stats_target['crmsd']
                crmsd[method] =  root_mean_square(models[method], observations[method])[1]
                rmsd[method] = stats_target['rmsd']

                if method == 0:
                    crmsdtay[method] = stats_taylor['crmsd'][0]
                    ccoef[method] = stats_taylor['ccoef'][0]
                    sdev[method] = stats_taylor['sdev'][0] / np.mean(observations[method])

                crmsdtay[method + 1] = stats_taylor['crmsd'][1]
                ccoef[method + 1] = stats_taylor['ccoef'][1]
                sdev[method + 1] = stats_taylor['sdev'][1] / np.mean(models[method])

                statistic_target_lake[method] = stats_target
                statistic_taylor_lake[method] = stats_taylor

            try:
                maxv = abs(round_decimals_up(max([max(bias, key=abs), max(crmsd, key=abs)])))
            except:
                maxv = 1
            try:
                minv = abs(round_decimals_down(min([min(bias, key=abs), min(crmsd, key=abs)])))
            except:
                minv = 0
            if not information == "all_lakes":
                sct_args = {'facecolor': colorpalette, 'marker': markers[lake],
                            'label': label_method, 'zorder': 100, 'alpha': 0.8, 's': sizes,
                            'edgecolors': 'k'}
                target_diagram_background(crmsd=crmsd, bias=bias, max_value_overall=maxv, min_value_overall=minv,
                                          ax=ax1, max_number_cicle=5,
                                          add_circles=True, sct_args=sct_args, cicle_max_args=circle_argsmax,
                                          cicle_min_args=circle_argsmin, cicle_1_args=circle_args,
                                          other_circle_args=other, color=colorpalette, label=label_method, size=sizes)

            else:
                if lake in speciallakes:
                    if variable == "Temperature":
                        colorpalettelake = [special[0],colorpalette[1]]
                        edge = ['k',special[0]]

                    else:
                        colorpalettelake = [special[1], colorpalette[1]]
                        edge = ['k', special[1]]

                else:
                    colorpalettelake = colorpalette
                    edge = ['k','k']
                sct_args = {'color': colorpalettelake, 'marker': markers[lake],
                            'label': label_method, 'zorder': 100, 'alpha': 0.8, 's': sizes,
                            'edgecolors': edge}
                biaslist[lake], crmsdlist[lake], rmsdlist[lake] = bias, crmsd, rmsd
                sdevlist[lake], crmsdtaylist[lake], ccoeflist[lake] = sdev, crmsdtay, ccoef


                print(len(all_data_from_model))
                if lake != (len(all_data_from_model) - 1):
                    target_diagram_background(crmsd=crmsd, bias=bias, max_value_overall=maxv, min_value_overall=minv,
                                              ax=ax1, max_number_cicle=5, add_circles=False, sct_args=sct_args,
                                              cicle_max_args=circle_argsmax,
                                              cicle_min_args=circle_argsmin, cicle_1_args=circle_args,
                                              other_circle_args=other, color=colorpalettelake, label=label_method,
                                              size=sizes,edgecolor=edge)
                else:
                    t1 = [max(i) for i in biaslist]
                    t2 = max(t1)
                    t3 = max([t2, max([max(i) for i in crmsdlist])])
                    try:
                        # maxv = abs(
                        # round_decimals_up(max([max([max(i) for i in biaslist]), max([max(i) for i in crmsdlist])])))
                        maxbias = [max(bias, key=abs) for bias in biaslist]
                        maxbiass = max(maxbias, key=abs)
                        maxcrmsd = [max(crmsd, key=abs) for crmsd in crmsdlist]
                        maxcrmsds = max(maxcrmsd, key=abs)
                        maxx = max([maxbiass, maxcrmsds], key=abs)
                        maxr = round_decimals_up(maxx)
                        maxv = abs(maxr)
                        maxv = 3.5
                        # maxv = abs(round_decimals_up( max([max([max(bias, key=abs)for bias in biaslist], key=abs),max([max(crmsd, key=abs) for crmsd in crmsdlist], key=abs)], key=abs)))
                    except:
                        maxv = 1
                    try:
                        minbias = [min(bias, key=abs) for bias in biaslist]
                        minbiass = min(minbias, key=abs)
                        mincrmsd = [min(crmsd, key=abs) for crmsd in crmsdlist]
                        mincrmsds = min(mincrmsd, key=abs)
                        minx = min([minbiass, mincrmsds], key=abs)
                        minr = round_decimals_down(minx)
                        minv = abs(minr)

                        # minv = abs(
                        #     round_decimals_down(
                        #         min([min([min(i) for i in biaslist]), min([min(i) for i in crmsdlist])])))
                    except:
                        minv = 0


                    sct_args = {'color': colorpalette, 'marker': markers[lake],
                                'label': label_method[method], 'zorder': 100, 'alpha': 0.8, 's': sizes[method],
                                'edgecolors': edge}
                    target_diagram_background(crmsd=crmsd, bias=bias, max_value_overall=maxv, min_value_overall=minv,
                                              ax=ax1, max_number_cicle=5, add_circles=True, sct_args=sct_args,
                                              cicle_max_args=circle_argsmax,
                                              cicle_min_args=circle_argsmin, cicle_1_args=circle_args,
                                              other_circle_args=other, color=colorpalettelake, label=label_method,
                                              size=sizes,edgecolor=edge)
            if lake == 0:
                # plt.legend()
                if information == "all_lakes":
                    biaslist[lake], crmsdlist[lake], rmsdlist[lake] = bias, crmsd, rmsd
                    sdevlist[lake], crmsdtaylist[lake], ccoeflist[lake] = sdev, crmsdtay, ccoef

        # plt.show()
        fig = plt.gcf()
        fig.set_size_inches(width, height)

        plt.style.context('seaborn-paper')
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        # plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams["font.family"] = self.font_family

        plt.savefig(
            os.path.join(self.output_folder, "Comparative_target_%s_%s_%s%s.png" % (variable, information, self.time,self.font_family)))
        plt.savefig(
            os.path.join(self.output_folder, "Comparative_target_%s_%s_%s%s.svg" % (variable, information, self.time,self.font_family)))
        plt.savefig(
            os.path.join(self.output_folder,
                         "Comparative_target_%s_%s_%s%s.eps" % (variable, information, self.time, self.font_family)),format='eps')
        plt.close()
        plt.clf()



    def graphique_secchi(self,x, y, xerr, yerr, calibration=False,old=False):

        plt.style.context('seaborn-paper')
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.MEDIUM_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title

        # Esthetic Parameters
        colorpalette = sns.color_palette("dark", 10)
        linecolor = colorpalette[9]
        lineStart = 0
        lineEnd = 20
        if calibration:
            if old:
                errorbararg = {'fmt': 'o', 'color': 'green', 'markersize': 8, 'capsize': 20,
                               'linewidth': 4, 'elinewidth': 4}
            else:
                errorbararg = {'fmt': 'o', 'color': 'darkolivegreen', 'markersize': 8, 'capsize': 20,
                               'linewidth': 4, 'elinewidth': 4}
        else:
            errorbararg = {'fmt': 'o', 'color': 'lime', 'markersize': 8, 'capsize': 20,
                       'linewidth': 4, 'elinewidth': 4}


        sns.set(font_scale=2)
        sns.set_style("ticks")
        plt.grid(False)


        #Figure
        fig, ax = plt.subplots(figsize=(15.0, 14.5))
        base_plot_comparison(x,y,lineStart=lineStart,lineEnd=lineEnd,ax=ax,linecolor=linecolor,font_family=self.font_family)
        error_bar_plot(x,y,xerr,yerr,errorbararg=errorbararg,ax=ax,SIZE=self.BIGGER_SIZE,font_family=self.font_family)

        plt.xlabel("Average Observed Secchi_Depth (m)")
        plt.ylabel("Average Modeled Secchi Depth (m)")
        plt.xlim(0, 15)
        plt.ylim(0, 20)
        fig.suptitle("")
        fig.tight_layout(pad=2)
        # ax.legend(loc='best')

        if calibration:
            fig.savefig('Secchi_mean_comparison_calibration_old_%s_%s.png' % (old,self.time), dpi=125)
        else:
            fig.savefig('Secchi_mean_comparison_old_%s_%s.png' % (old,self.time), dpi=125)
        plt.close()

    def graphiqueTO(self,x, y, z, symbol, variable, calibration=False,old=False, lakeid="",
                    outputfolder=r'F:\output'):

        # Esthetic Parameters:
        sns.set(font_scale=2)
        sns.set_style("ticks")
        plt.grid(False)
        colorpalette = sns.color_palette("dark", 10)

        print(len(x), len(y), len(z))


        #Arrange data for multiple lakes analysed
        if lakeid == "":
            xall, yall = [item for sublist in x for item in sublist], [item for sublist in y for item in sublist]
        else:
            xall,yall  = x,y


        lineStart = -1
        lineEnd = 35

        #Figure
        fig, ax = plt.subplots(figsize=(15.0, 12))
        if calibration:
            if old:
                ccmap = 'seismic_r'
            else:
                ccmap = 'twilight_shifted_r'
        else:
            ccmap = 'coolwarm_r'
        if variable == "Temperature (°C)":
            linearg={'color':colorpalette[0], 'label':"y= x", 'linewidth':4,'linestyle':'--'}
            base_plot_comparison(xall, yall,lineStart=lineStart,lineEnd=lineEnd, ax=None, linecolor="k",bigger=self.BIGGER_SIZE,font_family=self.font_family)
            line_plot(lineStart=lineStart,lineEnd=lineEnd,ax=ax,linearg=linearg,font_family=self.font_family)
            # ccmap = 'Blues'
            markers = ['s']*12
        else:
            linearg = { 'color':colorpalette[3], 'label':"y= x", 'linestyle':'--'}
            base_plot_comparison(xall, yall, lineStart=lineStart, lineEnd=lineEnd, ax=None, linecolor="k",bigger=self.BIGGER_SIZE,font_family=self.font_family)
            line_plot(lineStart=lineStart, lineEnd=lineEnd, ax=ax, linearg=linearg,font_family=self.font_family)
            # ccmap = 'Reds'
            markers = ['o']*12


        # markers = ["o", "v", "^", "s", "P", "*", ">", "X", "D", "<", "p", "d"]
        if lakeid == "":
            for i, c in enumerate(np.unique(symbol)):
                try:
                    cs = plt.scatter(x[i], y[i], c=z[i], marker=markers[c], s=90, cmap=ccmap, linewidths=1, edgecolors='k',
                                     alpha=0.8)
                except:
                    print("error in")
        else:
            try:
                i = int(symbol[0])
            except:
                i = 1
            print(i)
            # if i > 11:
            #     print("here")
            print(markers[i])
            cs = plt.scatter(xall, yall, c=z, marker=markers[i], s=90, cmap=ccmap, linewidths=1, edgecolors='k',
                             alpha=0.8)

        cb = plt.colorbar(cs)
        plt.clim(0.0, 1.0)
        cb.ax.tick_params(labelsize=14)

        cb.ax.invert_yaxis()

        fig.suptitle("")
        fig.tight_layout(pad=2)

        plt.xlabel("Observed %s" % variable,labelsize=self.BIGGER_SIZE)
        plt.ylabel("Modeled %s" % variable,labelsize=self.BIGGER_SIZE)
        plt.ylim(labelsize=self.BIGGER_SIZE)
        plt.xlim(labelsize=self.BIGGER_SIZE)
        plt.rcParams['font.size'] = '50'
        # ax.legend(loc='best')  # 'upper left')
        if calibration:
            if variable == "Temperature (°C)":
                fig.savefig(os.path.join(outputfolder, 'Temperature_comparison_calibrated_old_%s_%s_%s.png' % (old,lakeid,self.time)), dpi=125)

            else:
                fig.savefig(os.path.join(outputfolder, 'Oxygen_comparison_calibrated_old_%s_%s_%s.png' % (old,lakeid, self.time)),
                            dpi=125)
        else:
            if variable == "Temperature (°C)":
                fig.savefig(os.path.join(outputfolder, 'Temperature_comparison_old_%s_%s_%s.png' % (old,lakeid,self.time)),
                            dpi=125)
            else:
                fig.savefig(os.path.join(outputfolder, 'Oxygen_comparison_old_%s_%s_%s.png' % (old,lakeid, self.time)),
                            dpi=125)
        plt.close()

    def graphique_secchi_v2(self, x, y, xerr, yerr, calibration=False, old=False):

        # Esthetic Parameters
        colorpalette = sns.color_palette("dark", 10)
        sns.set_style("ticks", {"xtick.major.size": 100, "ytick.major.size": 100})
        plt.xticks(rotation=15)
        plt.rcParams.update({"font.family": self.font_family})
        font = {'family': self.font_family}


        plt.rcParams.update({"font.family": self.font_family})
        fig = plt.gcf()
        fig.set_size_inches(self.width, self.height)
        plt.rcParams["font.family"] = self.font_family
        linecolor = "black"#colorpalette[9]
        lineStart = 0
        lineEnd = 15

        if calibration:
            if old:
                errorbararg = {'fmt': 'o', 'color': 'green', 'markersize': 4, 'capsize':5,
                               'linewidth': 0.5, 'elinewidth':1}
            else:
                errorbararg = {'fmt': 'o', 'color': "#003300", 'markersize': 4, 'capsize': 5,
                               'linewidth': 0.5, 'elinewidth':1}
        else:
            errorbararg = {'fmt': 'o', 'color': "#006600",  'markersize': 4, 'capsize':5,
                               'linewidth': 0.5, 'elinewidth':1}

        # sns.set(font_scale=2)
        # sns.set_style("ticks")
        # plt.grid(False)

        # Figure
        fig, ax = plt.subplots(figsize=(15.0, 12))
        plt.xlim(0, 15)
        plt.ylim(0, 15)
        base_plot_comparison(x, y, lineStart=lineStart, lineEnd=lineEnd, ax=ax, linecolor=linecolor,bigger=self.BIGGER_SIZE,font_family=self.font_family)
        error_bar_plot(x, y, xerr, yerr, errorbararg=errorbararg, ax=ax,SIZE=self.BIGGER_SIZE,font_family=self.font_family)

        plt.xlabel("Average Observed Secchi Depth (m)",fontsize=self.BIGGER_SIZE)
        plt.ylabel("Average Modeled Secchi Depth (m)",fontsize=self.BIGGER_SIZE)
        plt.rcParams.update({"font.family": self.font_family})
        plt.xlim(0, 15)
        plt.ylim(0, 15)
        plt.xticks(np.arange(0, 15, 2))
        plt.yticks(np.arange(0, 15, 2))
        ax.set_xticklabels([0,2,4,6,8,10,12,14], va='bottom', ha='center')
        ax.set_yticklabels([" 0"," 2"," 4"," 6"," 8","10","12","14"], va= 'center', ha='left')
        fig.suptitle("")
        # fig.tight_layout(pad=2)
        # ax.legend(loc='best')


        # plt.style.context('seaborn-paper')
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        # plt.tight_layout(rect=[5,5, 1, 1])
        fig = plt.gcf()
        fig.set_size_inches(self.width, self.height)
        plt.rcParams["font.family"] = self.font_family

        plt.xlabel("Average Observed Secchi Depth (m)", fontsize=self.BIGGER_SIZE)
        plt.ylabel("Average Modeled Secchi Depth (m)", fontsize=self.BIGGER_SIZE)
        plt.rcParams.update({"font.family": self.font_family})
        plt.tight_layout()
        fig.set_size_inches(self.width, self.height)

        # plt.show()
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams["font.family"] = self.font_family


        if calibration:
            fig.savefig(
            os.path.join(self.output_folder,'Secchi_mean_comparison_calibration_old_%s_%s%s.png' % (old, self.time,self.font_family)), dpi=125)
            fig.savefig(
                os.path.join(self.output_folder, 'Secchi_mean_comparison_calibration_old_%s_%s%s.svg' % (old, self.time,self.font_family)),
                             dpi=125)
            fig.savefig(
                os.path.join(self.output_folder,
                             'Secchi_mean_comparison_calibration_old_%s_%s%s.eps' % (old, self.time, self.font_family)),
                dpi=125,format='eps')
        else:
            fig.savefig(
            os.path.join(self.output_folder,'Secchi_mean_comparison_old_%s_%s%s.png' % (old, self.time,self.font_family)), dpi=125)
            fig.savefig(
                os.path.join(self.output_folder, 'Secchi_mean_comparison_old_%s_%s%s.svg' % (old, self.time,self.font_family)), dpi=125)
            fig.savefig(
                os.path.join(self.output_folder,
                             'Secchi_mean_comparison_old_%s_%s%s.eps' % (old, self.time, self.font_family)), dpi=125,format='eps')

        plt.close()
        plt.clf()

    def graphiqueTO_v2(self, x, y, z, symbol, variable, calibration=False, old=False, lakeid="",
                    outputfolder=r'F:\output'):

        # Esthetic Parameters
        colorpalette = sns.color_palette("dark", 10)
        sns.set_style("ticks", {"xtick.major.size": 100, "ytick.major.size": 100})
        plt.xticks(rotation=15)
        plt.rcParams.update({"font.family": self.font_family})
        # font = {'family': self.font_family}

        plt.rcParams.update({"font.family": self.font_family})
        fig = plt.gcf()
        fig.set_size_inches(self.width, self.height)
        plt.rcParams["font.family"] = self.font_family
        linecolor = "black"  # colorpalette[9]
        lineStart = -0
        lineEnd = 30


        print(len(x), len(y), len(z))

        # Arrange data for multiple lakes analysed
        if lakeid == "":
            xall, yall = [item for sublist in x for item in sublist], [item for sublist in y for item in sublist]
        else:
            xall, yall = x, y

        # Figure
        fig, ax = plt.subplots(figsize=(15.0, 12))
        plt.xlim(0, 30)
        plt.ylim(0, 30)
        if calibration:
            if old:
                edge = 'twilight_shifted_r'

                ccmap = 'twilight_shifted_r'
                ccmap = 'seismic_r'

            else:
                edge = 'seismic_r'
                ccmap = 'seismic_r'


        else:
            ccmap = 'coolwarm_r'
            ccmap = 'seismic_r'




        if variable == "Temperature (°C)":
            linearg = {'color': colorpalette[0], 'label': "y= x", 'linewidth': 1, 'linestyle': '--'}
            base_plot_comparison(xall, yall, lineStart=lineStart, lineEnd=lineEnd, ax=ax, linecolor="k",bigger=self.BIGGER_SIZE,font_family=self.font_family)
            line_plot(lineStart=lineStart, lineEnd=lineEnd, ax=ax, linearg=linearg,font_family=self.font_family)
            # ccmap = 'Blues'
            markers = ['s'] * 12
        else:
            linearg = {'color': colorpalette[0], 'label': "y= x", 'linewidth': 1, 'linestyle': '--'}
            base_plot_comparison(xall, yall, lineStart=lineStart, lineEnd=lineEnd, ax=None, linecolor="k", bigger=self.BIGGER_SIZE,font_family=self.font_family)
            line_plot(lineStart=lineStart, lineEnd=lineEnd, ax=ax, linearg=linearg,font_family=self.font_family)
            # ccmap = 'Blues'
            markers = ['o'] * 12

        # markers = ["o", "v", "^", "s", "P", "*", ">", "X", "D", "<", "p", "d"]
        if lakeid == "":
            for i, c in enumerate(np.unique(symbol)):
                if 1==1:#try:
                    cs = plt.scatter(x[i], y[i], c=z[i], marker=markers[c], s=30, cmap=ccmap, linewidths=1,
                                     edgecolors='k', alpha=0.8)
                    if calibration:
                        plt.scatter(x[i], y[i], c=z[i], marker=markers[c], s=30, cmap=ccmap, linewidths=1,edgecolors='k',alpha=0.8)
                    else:
                        plt.scatter(x[i], y[i], c='white', marker=markers[c], s=30,  linewidths=1, edgecolors=plt.cm.get_cmap(ccmap)(z[i]), alpha=0.8)
                # except:
                #     print("error in")
        else:
            try:
                i = int(symbol[0])
            except:
                i = 1
            print(i)
            # if i > 11:
            #     print("here")
            print(markers[i])
            cs = plt.scatter(xall, yall, c=z, marker=markers[i], s=30, cmap=ccmap, linewidths=1, edgecolors='k',
                             alpha=0.8)



        plt.xlabel("Observed %s" % variable, fontsize=self.BIGGER_SIZE)
        plt.ylabel("Modeled %s" % variable, fontsize=self.BIGGER_SIZE)
        plt.rcParams.update({"font.family": self.font_family})
        plt.ylim(0,30)
        plt.xlim(0,30)
        plt.xticks(np.arange(0, 30, 5))
        plt.yticks(np.arange(0, 30,5))
        ax.set_xticklabels([0, 5, 10, 15, 20, 25, 30], va='bottom', ha='center')
        ax.set_yticklabels([" 0 ", " 5 ", "10 ", "15 ", "20 ", "25 ", "30 "], va='center', ha='left')
        fig.suptitle("")
        # fig.tight_layout(pad=2)
        # ax.legend(loc='best')
        # ax.tick_params(labelsize=self.BIGGER_SIZE)

        # plt.style.context('seaborn-paper')
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        # plt.tight_layout(rect=[5,5, 1, 1])
        fig = plt.gcf()
        fig.set_size_inches(self.width, self.height)
        plt.rcParams["font.family"] = self.font_family

        plt.xlabel("Observed %s" % variable, fontsize=self.BIGGER_SIZE)
        plt.ylabel("Modeled %s" % variable, fontsize=self.BIGGER_SIZE)
        plt.rcParams.update({"font.family": self.font_family})
        plt.tight_layout()
        fig.set_size_inches(self.width, self.height)

        # plt.show()
        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams["font.family"] = self.font_family



        # ax.legend(loc='best')  # 'upper left')
        if calibration:
            if variable == "Temperature (°C)":
                fig.savefig(os.path.join(outputfolder, 'Temperature_comparison_calibrated_old_%s_%s_%s%s.png' % (
                old, lakeid, self.time,self.font_family)), dpi=125)
                fig.savefig(os.path.join(outputfolder, 'Temperature_comparison_calibrated_old_%s_%s_%s%s.svg' % (
                    old, lakeid, self.time,self.font_family)), dpi=125)
                fig.savefig(os.path.join(outputfolder, 'Temperature_comparison_calibrated_old_%s_%s_%s%s.eps' % (
                    old, lakeid, self.time, self.font_family)), dpi=125,format='eps')

            else:
                fig.savefig(os.path.join(outputfolder,
                                         'Oxygen_comparison_calibrated_old_%s_%s_%s%s.png' % (old, lakeid, self.time,self.font_family)),
                            dpi=125)
                fig.savefig(os.path.join(outputfolder,
                                         'Oxygen_comparison_calibrated_old_%s_%s_%s%s.svg' % (old, lakeid, self.time,self.font_family)),
                            dpi=125)
                fig.savefig(os.path.join(outputfolder,
                                         'Oxygen_comparison_calibrated_old_%s_%s_%s%s.eps' % (
                                         old, lakeid, self.time, self.font_family)),
                            dpi=125,format='eps')
        else:
            if variable == "Temperature (°C)":
                fig.savefig(
                    os.path.join(outputfolder, 'Temperature_comparison_old_%s_%s_%s%s.png' % (old, lakeid, self.time,self.font_family)),
                    dpi=125)
                fig.savefig(
                    os.path.join(outputfolder, 'Temperature_comparison_old_%s_%s_%s%s.svg' % (old, lakeid, self.time,self.font_family)),
                    dpi=125)
                fig.savefig(
                    os.path.join(outputfolder, 'Temperature_comparison_old_%s_%s_%s%s.eps' % (
                    old, lakeid, self.time, self.font_family)),
                    dpi=125,format='eps')
            else:
                fig.savefig(os.path.join(outputfolder, 'Oxygen_comparison_old_%s_%s_%s%s.png' % (old, lakeid, self.time,self.font_family)),
                            dpi=125)
                fig.savefig(os.path.join(outputfolder, 'Oxygen_comparison_old_%s_%s_%s%s.svg' % (old, lakeid, self.time,self.font_family)),
                            dpi=125)
                fig.savefig(os.path.join(outputfolder, 'Oxygen_comparison_old_%s_%s_%s%s.eps' % (
                old, lakeid, self.time, self.font_family)),
                            dpi=125,format='eps')

        plt.close()
        plt.clf()

    def contourplot_temp_vs_light_oxy(self, x_list, y_list, z_list, variables_list_in_order, label_axis_in_order,
                                      subfolder="T_L_O_matrices",
                                      lakes_list=r'C:\Users\macot620\Documents\GitHub\Fish_niche\lakes\2017SwedenList.csv',
                                      individual=False):

        # lakes_list = r'C:\Users\macot620\Documents\GitHub\Fish_niche\lakes\2017SwedenList_only_validation_12lakessurface.csv'
        self.time = 2

        self.MINUS_SIZE = self.SMALL_SIZE
        plt.style.context('seaborn-paper')

        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.SMALL_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.MINUS_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.MINUS_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.MEDIUM_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams['axes.facecolor'] = 'silver'
        left = 0.10  # the left side of the subplots of the figure
        right = 0.80  # the right side of the subplots of the figure
        bottom = 0.03  # the bottom of the subplots of the figure
        top = 0.95  # the top of the subplots of the figure
        wspace = 0.2  # the amount of width reserved for blank space between subplots
        hspace = 0.5  # the amount of height reserved for white space between subplots

        # raw data and lake information
        lakes_data = pd.read_csv(lakes_list, encoding='ISO-8859-1')
        lake_id_list = list(lakes_data['lake_id'])

        [X, Y] = np.meshgrid(x_list, y_list)
        x_list = [float("{:.1f}".format(x_value)) for x_value in x_list]
        y_list = [float("{:.1f}".format(y_value)) for y_value in y_list]
        z_list = [float("{:.1f}".format(z_value)) for z_value in z_list]
        dict_all_z_value_by_lake_and_by_column = {}
        dict_all_z_value_by_lake = {}
        dict_z_value_by_lake, vmin_z_value_by_lake, vmax_z_value_by_lake = {}, {}, {}
        for z_value in z_list:
            dict_y_value_by_lake = {}
            for y_value in y_list:
                dict_raw_data_by_lake = {}
                for x_value in x_list:
                    dict_raw_data_by_lake, list_of_parameter_measured = get_file_by_value_of_xyz(x_value, y_value,
                                                                                                 z_value,
                                                                                                 dict_raw_data_by_lake,
                                                                                                 variables_list_in_order,
                                                                                                 os.path.join(
                                                                                                     self.output_folder,
                                                                                                     subfolder),
                                                                                                 lake_id_list)

                for lake in lake_id_list:
                    for column in range(0, len(list_of_parameter_measured)):
                        if not list_of_parameter_measured[column] in dict_y_value_by_lake:
                            dict_y_value_by_lake[list_of_parameter_measured[column]] = {
                                "%s" % lake: [[item[column] for item in dict_raw_data_by_lake["%s" % lake]]]}
                        else:
                            if not "%s" % lake in dict_y_value_by_lake[list_of_parameter_measured[column]]:
                                dict_y_value_by_lake[list_of_parameter_measured[column]]["%s" % lake] = [
                                    [item[column] for item in dict_raw_data_by_lake["%s" % lake]]]

                            else:
                                dict_y_value_by_lake[list_of_parameter_measured[column]]["%s" % lake].append(
                                    [item[column] for item in dict_raw_data_by_lake["%s" % lake]])

            dict_all_z_value_by_lake_and_by_column['%s' % z_value] = dict_y_value_by_lake

        dict_z_value_by_lake, vmin, vmax = create_dictionnary_of_z_value_by_lake(dict_all_z_value_by_lake_and_by_column,
                                                                                 z_list, list_of_parameter_measured,
                                                                                 lake_id_list, dict_z_value_by_lake)

        for column in list_of_parameter_measured:
            if column == "NTGdays":
                for lake in lake_id_list:
                    if individual:
                        for z_value in z_list:
                            fig, ax = plt.subplots(1, 1, figsize=(10, 8))
                            base_contourplot(X, Y, dict_z_value_by_lake[column]['%s' % lake]['%s' % z_value], [z_value],
                                             lake, 4,
                                             vmin[column]['%s' % lake], vmax[column]['%s' % lake], ax,
                                             individual=individual)
                            ax.set_ylabel(label_axis_in_order['ylabel'], fontsize=self.BIGGER_SIZE)
                            ax.set_xlabel(label_axis_in_order['xlabel'], fontsize=self.BIGGER_SIZE)
                            plt.savefig("%s/habitat_treshold/contourplot_%s_lake_%s_%s_%s_time%s.png" % (
                                self.output_folder, column, lake, variables_list_in_order[2], z_value, self.time))

                    fig, axs = plt.subplots(4, 4, figsize=(15, 15))
                    fig.set_size_inches(self.width, self.height)
                    print(vmax[column]['%s' % lake])
                    if column == "NTGdays":
                        listdata = dict_z_value_by_lake[column]['%s' % lake]

                        contour = base_contourplot(X, Y, listdata, z_list, lake, 4,
                                                   0, 365, axs, ntg=True)
                    else:
                        contour = base_contourplot(X, Y, dict_z_value_by_lake[column]['%s' % lake], z_list, lake, 4,
                                                   vmin[column]['%s' % lake], vmax[column]['%s' % lake], axs)

                    # axs[3][3].set_visible(False)
                    # plt.subplots_adjust(right=right, left=left,top=top, wspace=wspace, hspace=hspace)
                    # cbar_ax = fig.add_axes([0.85, 0.15, 0.05, 0.7])
                    print(np.linspace(vmin[column]['%s' % lake], vmax[column]['%s' % lake], 10))
                    # if column == "NTGdays":
                    #
                    #     clb = plt.colorbar(contour,  cax=cbar_ax,boundaries=np.linspace(0, 365, 10))
                    #     # plt.show()
                    #     # plt.clim(0, 365)
                    # else:
                    #     clb = plt.colorbar(contour, cax=cbar_ax,
                    #                    boundaries=np.linspace(vmin[column]['%s' % lake], vmax[column]['%s' % lake], 10))

                    # clb = plt.colorbar(contour, cax=cbar_ax)
                    # clb.set_label(column)

                    # plt.subplots_adjust(left, bottom, right, top, wspace, hspace)
                    fig.text(0.5, 0.02, label_axis_in_order['xlabel'], ha='center', va='center',
                             fontsize=self.MEDIUM_SIZE)
                    fig.text(0.02, 0.5, label_axis_in_order['ylabel'], ha='center', va='center', rotation='vertical',
                             fontsize=self.MEDIUM_SIZE)

                    plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
                    plt.rc('axes', titlesize=self.SMALL_SIZE)  # fontsize of the axes title
                    plt.rc('axes', labelsize=self.SMALL_SIZE)  # fontsize of the x and y labels
                    plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
                    plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
                    plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
                    plt.rc('figure', titlesize=self.SMALL_SIZE)  # fontsize of the figure title

                    # fig.subplots_adjust(hspace=0.3)
                    fig.set_size_inches(self.width, self.height)
                    #
                    fig.subplots_adjust(left=0.07, right=.95, top=0.95, bottom=0.07, hspace=0.4)
                    plt.tight_layout(h_pad=0.5, rect=[0.05, 0.05, 0.05, 0.05])
                    fig.set_size_inches(self.width, self.height)

                    plt.rcParams["font.family"] = self.font_family
                    plt.rcParams.update({"font.family": self.font_family})

                    print(r"G:\Fish_Niche_archive\Postproc\figures_surface_area_NTGdays/contourplot_%s_lake_%s_%s.png" % (
                        column, lake, variables_list_in_order[2],))
                    # plt.savefig(r"G:\Fish_Niche_archive\Postproc\figures_surface_area_NTGdays/contourplot_%s_lake_%s_%s.png" % (
                    #     column, lake, variables_list_in_order[2]))
                    # plt.savefig(r"G:\Fish_Niche_archive\Postproc\figures_surface_area_NTGdays/contourplot_%s_lake_%s_%s.svg" % (
                    #     column, lake, variables_list_in_order[2]))
                    plt.savefig(
                        r"G:\Fish_Niche_archive\Postproc\figures_surface_area_NTGdays/contourplot_%s_lake_%s_%s.jpeg" % (
                            column, lake, variables_list_in_order[2]))
                    plt.savefig(r"G:\Fish_Niche_archive\Postproc\figures_surface_area_NTGdays/contourplot_%s_lake_%s_%s.eps" % (
                        column, lake, variables_list_in_order[2]),format='eps')
                    plt.close('all')

    def contourplot_temp_vs_light_oxy_trueversion(self,x_list,y_list,z_list,variables_list_in_order,label_axis_in_order,
                                      subfolder = "T_L_O_matrices",lakes_list=r'C:\Users\macot620\Documents\GitHub\Fish_niche\lakes\2017SwedenList.csv',individual=False):

        #lakes_list = r'C:\Users\macot620\Documents\GitHub\Fish_niche\lakes\2017SwedenList_only_validation_12lakessurface.csv'
        self.time = 2

        self.MINUS_SIZE = self.SMALL_SIZE
        plt.style.context('seaborn-paper')

        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.SMALL_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.MINUS_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.MINUS_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.MEDIUM_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams['axes.facecolor'] = 'silver'
        left = 0.10  # the left side of the subplots of the figure
        right = 0.80  # the right side of the subplots of the figure
        bottom = 0.03  # the bottom of the subplots of the figure
        top = 0.95  # the top of the subplots of the figure
        wspace = 0.2  # the amount of width reserved for blank space between subplots
        hspace = 0.5  # the amount of height reserved for white space between subplots

        # raw data and lake information
        lakes_data = pd.read_csv(lakes_list, encoding='ISO-8859-1')
        lake_id_list = list(lakes_data['lake_id'])


        [X, Y] = np.meshgrid(x_list, y_list)
        x_list = [float("{:.1f}".format(x_value)) for x_value in x_list]
        y_list = [float("{:.1f}".format(y_value)) for y_value in y_list]
        z_list = [float("{:.1f}".format(z_value)) for z_value in z_list]
        dict_all_z_value_by_lake_and_by_column = {}
        dict_all_z_value_by_lake = {}
        dict_z_value_by_lake, vmin_z_value_by_lake, vmax_z_value_by_lake = {}, {}, {}
        for z_value in z_list:
            dict_y_value_by_lake = {}
            for y_value in y_list:
                dict_raw_data_by_lake = {}
                for x_value in x_list:
                    dict_raw_data_by_lake,list_of_parameter_measured = get_file_by_value_of_xyz(x_value, y_value, z_value, dict_raw_data_by_lake,
                                                                variables_list_in_order, os.path.join(self.output_folder, subfolder), lake_id_list)

                for lake in lake_id_list:
                    for column in range(0, len(list_of_parameter_measured)):
                        if not list_of_parameter_measured[column] in dict_y_value_by_lake:
                            dict_y_value_by_lake[list_of_parameter_measured[column]]={"%s" % lake:[[item[column] for item in  dict_raw_data_by_lake["%s" % lake]]]}
                        else:
                            if not "%s"%lake in dict_y_value_by_lake[list_of_parameter_measured[column]]:
                                dict_y_value_by_lake[list_of_parameter_measured[column]]["%s" % lake] =  [[item[column] for item in dict_raw_data_by_lake["%s" % lake]]]

                            else:
                                dict_y_value_by_lake[list_of_parameter_measured[column]]["%s" % lake].append([item[column] for item in dict_raw_data_by_lake["%s" % lake]])




            dict_all_z_value_by_lake_and_by_column['%s'%z_value] = dict_y_value_by_lake



        dict_z_value_by_lake,vmin,vmax = create_dictionnary_of_z_value_by_lake(dict_all_z_value_by_lake_and_by_column,z_list,list_of_parameter_measured,lake_id_list,dict_z_value_by_lake)


        for column in list_of_parameter_measured:
            if column == "NTGdays":
                for lake in lake_id_list:
                    if individual:
                        for z_value in z_list:
                            fig, ax = plt.subplots(1, 1, figsize=(10, 8))
                            base_contourplot(X,Y,dict_z_value_by_lake[column]['%s'%lake]['%s'%z_value], [z_value],lake, 4,
                                             vmin[column]['%s'%lake],vmax[column]['%s'%lake],ax,individual=individual)
                            ax.set_ylabel(label_axis_in_order['ylabel'],fontsize=self.BIGGER_SIZE)
                            ax.set_xlabel(label_axis_in_order['xlabel'],fontsize=self.BIGGER_SIZE)
                            plt.savefig("%s/habitat_treshold/contourplot_%s_lake_%s_%s_%s_time%s.png" % (
                                self.output_folder, column, lake,variables_list_in_order[2],z_value, self.time))


                    fig, axs = plt.subplots(4, 4, figsize=(15, 15))
                    fig.set_size_inches(self.width, self.height)
                    print(vmax[column]['%s'%lake])
                    if column == "NTGdays":
                        listdata = dict_z_value_by_lake[column]['%s'%lake]

                        contour = base_contourplot(X, Y, listdata, z_list, lake, 4,
                                         0, 365, axs, ntg=True)
                    else:
                        contour = base_contourplot(X, Y, dict_z_value_by_lake[column]['%s' % lake], z_list, lake, 4,
                                                   vmin[column]['%s' % lake], vmax[column]['%s' % lake], axs)

                    # axs[3][3].set_visible(False)
                    # plt.subplots_adjust(right=right, left=left,top=top, wspace=wspace, hspace=hspace)
                    # cbar_ax = fig.add_axes([0.85, 0.15, 0.05, 0.7])
                    print(np.linspace(vmin[column]['%s'%lake], vmax[column]['%s'%lake], 10))
                    # if column == "NTGdays":
                    #
                    #     clb = plt.colorbar(contour,  cax=cbar_ax,boundaries=np.linspace(0, 365, 10))
                    #     # plt.show()
                    #     # plt.clim(0, 365)
                    # else:
                    #     clb = plt.colorbar(contour, cax=cbar_ax,
                    #                    boundaries=np.linspace(vmin[column]['%s' % lake], vmax[column]['%s' % lake], 10))

                    # clb = plt.colorbar(contour, cax=cbar_ax)
                    # clb.set_label(column)

                    # plt.subplots_adjust(left, bottom, right, top, wspace, hspace)
                    fig.text(0.5, 0.02, label_axis_in_order['xlabel'], ha='center', va='center',fontsize=self.MEDIUM_SIZE)
                    fig.text(0.02, 0.5, label_axis_in_order['ylabel'], ha='center', va='center', rotation='vertical',fontsize=self.MEDIUM_SIZE)

                    plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
                    plt.rc('axes', titlesize=self.SMALL_SIZE)  # fontsize of the axes title
                    plt.rc('axes', labelsize=self.SMALL_SIZE)  # fontsize of the x and y labels
                    plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
                    plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
                    plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
                    plt.rc('figure', titlesize=self.SMALL_SIZE)  # fontsize of the figure title

                    # fig.subplots_adjust(hspace=0.3)
                    fig.set_size_inches(self.width, self.height)
                    #
                    fig.subplots_adjust(left=0.07, right=.95,top=0.95,bottom=0.07, hspace=0.4)
                    plt.tight_layout(h_pad=0.5, rect=[0.05, 0.05, 0.05, 0.05])
                    fig.set_size_inches(self.width, self.height)

                    plt.rcParams["font.family"] = self.font_family
                    plt.rcParams.update({"font.family": self.font_family})


                    print("%s/habitat_treshold3/contourplot_%s_lake_%s_%s_time%s.png" % (
                                self.output_folder, column, lake,variables_list_in_order[2], self.time))
                    plt.savefig("%s/habitat_treshold3/contourplot_%s_lake_%s_%s_time%s.png" % (
                                self.output_folder, column, lake,variables_list_in_order[2], self.time))
                    plt.savefig("%s/habitat_treshold3/contourplot_%s_lake_%s_%s_time%s.svg" % (
                        self.output_folder, column, lake, variables_list_in_order[2], self.time))
                    plt.savefig("%s/habitat_treshold3/contourplot_%s_lake_%s_%s_time%s.eps" % (
                        self.output_folder, column, lake, variables_list_in_order[2], self.time), format="eps")
                    plt.close('all')

    def timeseries_habitat_by_model(self,listmodels=[2],  partype="50",uninhabitable=True):
        years = YearLocator()  # every year
        months = MonthLocator()  # every month
        yearsFmt = DateFormatter('%Y')
        monthsFmt = DateFormatter('%M')
        # fig2,fig3=plt.figure(),plt.figure()
        modelname = ['KNM', 'DMI', 'MPI', 'MOH', 'IPS', 'CNR']
        color2 = ['white', 'blue', 'black', 'magenta', 'cyan', 'red', 'yellow']
        scenarios=['historical', 'rcp45', 'rcp85']
        i = 0
        for group in [1, 2, 3]:

            datasheet_all = pd.read_csv(os.path.join(self.output_folder, 'complete_data_%s%s.csv' % (group, partype)))
            datasheet_all.columns = ["Date", "Model", "Scenario", "Lake_group", "Lake_id", "pT", "pO2", "pPAR",
                                     "TotalVolume", "habitable"]
            datasheet_all['Date'] = pd.to_datetime(datasheet_all['Date'], format="%Y-%m-%d")
            datasheet_all.set_index('Date', inplace=True)
            datasheet2 = datasheet_all
            subplot = -1
            fig1, axes = plt.subplots(len(scenarios), 1, gridspec_kw={
                'width_ratios': [3],
                'height_ratios': [1] * len(scenarios)}, figsize=(10.0, 30.0))

            for scenario in scenarios:
                i += 1
                subplot += 1


                tt = pd.date_range(start='2000-01-01', end='2000-12-31')

                datasheet_scenario = datasheet_all.loc[datasheet_all['Scenario'] == scenario]
                if scenario == 'historical':
                    datasheet_scenario = datasheet_scenario.loc[:'1982-12-31']
                    datasheet_scenario['Scenario'] = 2
                else:
                    datasheet_scenario = datasheet_scenario.loc['2090-12-31':]
                    datasheet_scenario['Scenario'] = 8
                    # datasheet_scenario = datasheet_scenario.loc['2060-12-31':'2080-12-31']
                    # datasheet_scenario = datasheet_scenario.loc[:'2050-12-31']
                # ttyear = datasheet2.index
                # ttyear = ttyear.drop_duplicates(keep='first')
                listonplot = []
                # fig1 = plt.figure((2 * 100) + (group * 10) + (i), figsize=(18.0, 10.0))


                for model in listmodels:
                    datasheet = datasheet_scenario.loc[datasheet_scenario['Model'] == model]
                    if uninhabitable:
                        datasheet['p0T'] = 100-datasheet['pT'] * 100
                        datasheet['p0O2'] = 100-datasheet['pO2'] * 100
                        datasheet['p0PAR'] = 100-datasheet['pPAR'] * 100
                        datasheet['p0habitable'] = datasheet['habitable'] * 100
                    else:
                        datasheet['p0T'] = datasheet['pT'] * 100
                        datasheet['p0O2'] = datasheet['pO2'] * 100
                        datasheet['p0PAR'] = datasheet['pPAR'] * 100
                        datasheet['p0habitable'] = datasheet['habitable'] * 100

                    if len(datasheet) != 0:
                        listonplot.append(model)
                        medianbyday = datasheet.groupby([datasheet.index.month, datasheet.index.day]).quantile(0.5)
                        minbyday = datasheet.groupby([datasheet.index.month, datasheet.index.day]).quantile(0.25)
                        maxbyday = datasheet.groupby([datasheet.index.month, datasheet.index.day]).quantile(0.75)

                        meanbyday = datasheet.groupby([datasheet.index.month, datasheet.index.day]).mean()
                        stdbyday = datasheet.groupby([datasheet.index.month, datasheet.index.day]).std()

                        # z_critical = stats.norm.ppf(q=0.975) # Get the z-critical value*
                        meanbyday['p0O22'] = meanbyday['p0O2']
                        meanbyday['p0PAR2'] = meanbyday['p0PAR']
                        stdbyday['p0O22'] = stdbyday['p0O2']
                        stdbyday['p0PAR2'] = stdbyday['p0PAR']
                        # minbyday['p0O22'] =100- minbyday['p0O2']
                        # maxbyday['p0O22'] = 100-maxbyday['p0O2']
                        # minbyday['p0PAR2'] = 100 - minbyday['p0PAR']
                        # maxbyday['p0PAR2'] = 100 - maxbyday['p0PAR']
                        datahabitable = pd.DataFrame()
                        # datahabitable["Date"]=tt
                        datahabitable["min"] = minbyday['p0habitable']
                        datahabitable["max"] = maxbyday['p0habitable']
                        datahabitable["median"] = medianbyday['p0habitable']
                        datahabitable["mean"] = meanbyday['p0habitable']
                        datahabitable["std"] = stdbyday['p0habitable']
                        datahabitable["meanT"] = meanbyday['p0T']
                        datahabitable["meanO2"] = meanbyday['p0O2']
                        datahabitable["meanPAR"] = meanbyday['p0PAR']

                        print('hrer')
                        path_excel = os.path.join(self.output_folder, "habitable_all50_final.xlsx")
                        if not os.path.exists(path_excel):
                            workbook = xlsxwriter.Workbook(path_excel)
                            workbook.close()
                        book = load_workbook(path_excel)
                        writer = pd.ExcelWriter(path_excel, engine='openpyxl')
                        writer.book = book

                        datahabitable.to_excel(writer, sheet_name='g%s_s%s_m%s' % (group, scenario, model),
                                               index=False)
                        writer.save()
                        writer.close()

                        stats.norm.ppf(q=0.025)
                        # margin_of_error = z_critical *(stdbyday/sqrt(countbyday.iloc[0,0]))
                        meanplusmargin = meanbyday + stdbyday
                        meanplusmargin.loc[meanplusmargin['p0T'] > 100, 'p0T'] = 100
                        meanplusmargin.loc[meanplusmargin['p0PAR2'] > 100, 'p0PAR2'] = 100
                        meanplusmargin.loc[meanplusmargin['p0O22'] > 100, 'p0O22'] = 100
                        meanlessmargin = meanbyday - stdbyday
                        meanlessmargin.loc[meanlessmargin['p0T'] < 0, 'p0T'] = 0
                        meanlessmargin.loc[meanlessmargin['p0PAR2'] < 0, 'p0PAR2'] = 0
                        meanlessmargin.loc[meanlessmargin['p0O22'] < 0, 'p0O22'] = 0




                        axes[subplot].fill_between(tt, meanbyday['p0T'], meanlessmargin['p0T'], color='#2A4D60', alpha='0.2',zorder=100)
                        axes[subplot].fill_between(tt, meanbyday['p0T'], meanplusmargin['p0T'], color='#2A4D60', alpha='0.2',zorder=100)
                        # axes[subplot].plot_date(tt, meanbyday['p0T'], '-', color=plt.cm.binary((model / 10) + 0.4), lw=2, ms=3)

                        axes[subplot].fill_between(tt, meanbyday['p0O22'], meanlessmargin['p0O22'], color='#C00000', alpha='0.2',zorder=100)
                        axes[subplot].fill_between(tt, meanbyday['p0O22'], meanplusmargin['p0O22'], color='#C00000', alpha='0.2',zorder=100)
                        # axes[subplot].plot_date(tt, meanbyday['p0O22'], '-', color=plt.cm.binary((model / 10) + 0.4), lw=2,ms=3)

                        axes[subplot].fill_between(tt, meanbyday['p0PAR2'], meanplusmargin['p0PAR2'], color='#70AD47', alpha='0.2',zorder=100)
                        axes[subplot].fill_between(tt, meanbyday['p0PAR2'], meanlessmargin['p0PAR2'], color='#70AD47', alpha='0.2',zorder=100)
                        # axes[subplot].plot_date(tt, meanbyday['p0PAR2'], '-', color=plt.cm.binary((model / 10) + 0.4), lw=2, ms=3)

                        axes[subplot].plot_date(tt, meanbyday['p0T'], '--', color='#2A4D60', lw=2, ms=3, zorder=1)
                        axes[subplot].plot_date(tt, meanbyday['p0O22'], '--', color='#C00000', lw=2, ms=3, zorder=1)
                        axes[subplot].plot_date(tt, meanbyday['p0PAR2'], '--', color='#70AD47',lw=2, ms=3,zorder=1)

                        axes[subplot].autoscale_view()

                        axes[subplot].set_xlim([datetime(2000, 1, 1), datetime(2000, 12, 31)])
                        axes[subplot].set_ylim([0, 100])
                        axes[subplot].set_yticks([0, 25, 50, 75, 100])
                        # plt.yticks([0, 25, 50, 75, 100])

                        if subplot == 2:
                            axes[subplot].set(xlabel='Date',ylabel="% Total Habitable\n Volume")
                        else:
                            axes[subplot].set( ylabel="% Total Habitable\n Volume")
                        # plt.ylabel("% Volume by variable")

                        ax2 = axes[subplot].twinx()
                        ax2.plot_date(tt, datahabitable["mean"], '-', color="black", lw=2, ms=3)

                        ax2.set_ylim(0, 100, 50)
                        ax2.set_yticks([0, 25, 50, 75, 100])
                        # plt.yticks([0, 25, 50, 75, 100])

                        axes[subplot].xaxis.set_major_locator(MonthLocator())
                        axes[subplot].xaxis.set_minor_locator(mondays)
                        axes[subplot].xaxis.set_major_formatter(weekFormatter)
                        # ax1.fmt_ydata = price
                        axes[subplot].yaxis.grid(True,color='grey', linestyle='--', linewidth=0.5,alpha=0.5)
                        if uninhabitable:
                            ax2.set(ylabel="% Uninhabitable Volume\n by Variable (T,DO,L)")
                        else:
                            ax2.set(ylabel="% Habitable Volume\n by Variable (T,DO,L)")

            fig1.set_size_inches(self.width, self.height)
            plt.tight_layout()
            fig1.subplots_adjust(hspace=0.3)

            plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
            plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
            plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
            plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
            plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
            plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
            plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
            plt.rcParams["font.family"] = self.font_family
            if uninhabitable:

                fig1.savefig(
                    os.path.join(self.output_folder,
                              "Figure_synthese_variable_and_total_group_%s_scenario_%s_mean_%s_uninhabitable%s.svg" % (group, scenario, partype,self.font_family)))
                fig1.savefig( os.path.join(self.output_folder,
                         "Figure_synthese_variable_and_total_group_%s_scenario_%s_mean_%s_uninhabitable%s.eps" % (
                         group, scenario, partype, self.font_family)), format = "eps")
                fig1.savefig(
                    os.path.join(self.output_folder,
                              "Figure_synthese_ariable_and_total_group_%s_scenario_%s_mean_%s_uninhabitable%s.png" % (group, scenario, partype,self.font_family)))
            else:

                fig1.savefig(
                    os.path.join(self.output_folder,
                                 "Figure_synthese_variable_and_total_group_%s_scenario_%s_mean_%s_habitable%s.svg" % (
                                 group, scenario, partype,self.font_family)))
                fig1.savefig(
                    os.path.join(self.output_folder,
                                 "Figure_synthese_variable_and_total_group_%s_scenario_%s_mean_%s_habitable%s.eps" % (
                                     group, scenario, partype, self.font_family)), format = "eps")
                fig1.savefig(
                    os.path.join(self.output_folder,
                                 "Figure_synthese_ariable_and_total_group_%s_scenario_%s_mean_%s_habitable%s.png" % (
                                 group, scenario, partype,self.font_family)))
            print('completed')
            fig1.clear()
            plt.close()

    # def contourplot_temp_vs_oxy_light_old_version(self,temp_list,light_list,oxygen_list, subfolder = "T_L_O_matrices", lakes_list=r'C:\Users\macot620\Documents\GitHub\Fish_niche\lakes\2017SwedenList_only_validation_12lakes.csv',individual=False):
    #     #self.time = 2
    #     SMALL_SIZE = 12
    #     MEDIUM_SIZE = 14
    #     BIGGER_SIZE = 16
    #     plt.style.context('seaborn-paper')
    #     plt.rc('font', size=MEDIUM_SIZE)  # controls default text sizes
    #     plt.rc('axes', titlesize=MEDIUM_SIZE)  # fontsize of the axes title
    #     plt.rc('axes', labelsize=MEDIUM_SIZE)  # fontsize of the x and y labels
    #     plt.rc('xtick', labelsize=SMALL_SIZE)  # fontsize of the tick labels
    #     plt.rc('ytick', labelsize=SMALL_SIZE)  # fontsize of the tick labels
    #     plt.rc('legend', fontsize=SMALL_SIZE)  # legend fontsize
    #     plt.rc('figure', titlesize=MEDIUM_SIZE)  # fontsize of the figure title
    #
    #     lakes_data = pd.read_csv(lakes_list, encoding='ISO-8859-1')
    #     plt.rcParams['axes.facecolor'] = 'silver'
    #     AreaDays_o2, Areaday_o2, NTGdays_o2 = {}, {}, {}
    #     first3 = 0
    #     vminAD, vmaxAD = {},{}
    #     vminAd, vmaxAd ={},{}
    #     vminNTG, vmaxNTG = {},{}
    #     for light in light_list:
    #         if light != int(light):
    #             light = "%s0" % light
    #
    #         [X, Y] = np.meshgrid(oxygen_list, temp_list)
    #         first2 = 0
    #         AreaDays, Areaday, NTGdays = {}, {}, {}
    #         for temp in temp_list:
    #             temperatures = {}
    #             first = 0
    #             for oxygen in oxygen_list:
    #                 if oxygen != str(oxygen):
    #                     if oxygen != int(oxygen):
    #                         oxygen = "%s0" % oxygen
    #                 with open('%s/%s/fish_niche_Area_Light%s_T%s_O%s_2001-2010.csv' % (
    #                 self.output_folder, subfolder, light, temp, oxygen), newline='') as f:
    #                     reader = csv.reader(f)
    #                     data = list(reader)
    #
    #                 for lake in range(0, len(lakes_data["lake_id"])):
    #                     lake += 1
    #                     if first == 0:
    #                         if data[lake][1] == 'AreaDays':
    #                             temperatures["%s" % data[lake][0]] = [[data[lake][1], data[lake][2], data[lake][3]]]
    #                         else:
    #                             temperatures["%s" % data[lake][0]] = [
    #                                 [float(data[lake][1]), float(data[lake][2]), float(data[lake][3])]]
    #
    #
    #                     else:
    #                         if data[lake][1] == 'AreaDays':
    #                             temperatures["%s" % data[lake][0]].append([data[lake][1], data[lake][2], data[lake][3]])
    #                         else:
    #                             temperatures["%s" % data[lake][0]].append(
    #                                 [float(data[lake][1]), float(data[lake][2]), float(data[lake][3])])
    #
    #                 first = 1
    #
    #             for lake in range(0, len(lakes_data["lake_id"])):
    #                 lake += 1
    #                 if first2 == 0:
    #                     print(temperatures["%s" % data[lake][0]][0])
    #                     AreaDays["%s" % data[lake][0]] = [[item[0] for item in temperatures["%s" % data[lake][0]]]]
    #                     Areaday["%s" % data[lake][0]] = [[item[1] for item in temperatures["%s" % data[lake][0]]]]
    #                     NTGdays["%s" % data[lake][0]] = [[item[2] for item in temperatures["%s" % data[lake][0]]]]
    #
    #                     vminAD["%s" % data[lake][0]] = min([item[0] for item in temperatures["%s" % data[lake][0]]])
    #                     vmaxAD["%s" % data[lake][0]] = max([item[0] for item in temperatures["%s" % data[lake][0]]])
    #                     vminAd["%s" % data[lake][0]] = min([item[1] for item in temperatures["%s" % data[lake][0]]])
    #                     vmaxAd["%s" % data[lake][0]] = max([item[1] for item in temperatures["%s" % data[lake][0]]])
    #                     vminNTG["%s" % data[lake][0]]= min([item[2] for item in temperatures["%s" % data[lake][0]]])
    #                     vmaxNTG["%s" % data[lake][0]] = max([item[2] for item in temperatures["%s" % data[lake][0]]])
    #
    #                 else:
    #                     AreaDays["%s" % data[lake][0]].append([item[0] for item in temperatures["%s" % data[lake][0]]])
    #                     Areaday["%s" % data[lake][0]].append([item[1] for item in temperatures["%s" % data[lake][0]]])
    #                     NTGdays["%s" % data[lake][0]].append([item[2] for item in temperatures["%s" % data[lake][0]]])
    #                     if min([item[0] for item in temperatures["%s" % data[lake][0]]]) < vminAD["%s" % data[lake][0]]:
    #                         vminAD["%s" % data[lake][0]] = min([item[0] for item in temperatures["%s" % data[lake][0]]])
    #                     if max([item[0] for item in temperatures["%s" % data[lake][0]]]) > vmaxAd["%s" % data[lake][0]]:
    #                         vmaxAD["%s" % data[lake][0]] = max([item[0] for item in temperatures["%s" % data[lake][0]]])
    #                     if min([item[1] for item in temperatures["%s" % data[lake][0]]]) < vminAd["%s" % data[lake][0]]:
    #                         vminAd["%s" % data[lake][0]] = min([item[1] for item in temperatures["%s" % data[lake][0]]])
    #                     if max([item[1] for item in temperatures["%s" % data[lake][0]]]) > vmaxAd["%s" % data[lake][0]]:
    #                         vmaxAd["%s" % data[lake][0]] = max([item[1] for item in temperatures["%s" % data[lake][0]]])
    #                     if min([item[2] for item in temperatures["%s" % data[lake][0]]]) < vminNTG["%s" % data[lake][0]]:
    #                         vminNTG["%s" % data[lake][0]] = min([item[2] for item in temperatures["%s" % data[lake][0]]])
    #                     if max([item[2] for item in temperatures["%s" % data[lake][0]]]) > vmaxNTG["%s" % data[lake][0]]:
    #                         vmaxNTG["%s" % data[lake][0]] = max([item[2] for item in temperatures["%s" % data[lake][0]]])
    #             first2 = 1
    #
    #         if first3 == 0:
    #             AreaDays_o2["%s" % light] = AreaDays
    #             Areaday_o2["%s" % light] = Areaday
    #             NTGdays_o2["%s" % light] = NTGdays
    #
    #
    #         if individual:
    #             for lake in lakes_data["lake_id"]:
    #                 Z = AreaDays["%s" % lake]
    #                 fig, ax = plt.subplots(1, 1, figsize=(10, 8))
    #
    #                 cp1 =ax.contourf(X, Y, Z,vmin=vminAD["%s" % lake],vmax=vmaxAD["%s" % lake])
    #                 # fig.colorbar(cp)  # Add a colorbar to a plot
    #                 cp = plt.contour(X, Y, Z, colors='white')
    #                 fig.colorbar(cp1)
    #                 plt.clabel(cp, inline=True, fmt='%.2e')
    #                 plt.title("light threshold of %s µmol/m2s" % light)
    #                 # divider = make_axes_locatable(ax)
    #                 # cax = divider.append_axes("top", size="8%", pad=0)
    #                 # cax.get_xaxis().set_visible(False)
    #                 # cax.get_yaxis().set_visible(False)
    #                 # cax.set_facecolor('#e0e0e0')
    #                 #
    #                 # at = AnchoredText(("Oxygen threshold of %s mg/l")%oxygen, loc=10,
    #                 #                   prop=dict(backgroundcolor='#e0e0e0',color='#e0e0e0',size=12))
    #                 #
    #                 # cax.add_artist(at)
    #                 ax.set_ylabel('Temperature threshold (°C)')
    #                 ax.set_xlabel('Oxygen threshold (mg/L)')
    #                 plt.savefig("%s/habitat_treshold/contourplot_AreaDays_lake_%s_light_%s_time%s.png" % (
    #                 self.output_folder, lake, light, self.time))
    #
    #                 Z = Areaday["%s" % lake]
    #                 fig, ax = plt.subplots(1, 1, figsize=(10, 8))
    #                 cp1 = ax.contourf(X, Y, Z,vmin=vminAd["%s" % lake],vmax=vmaxAd["%s" % lake])
    #                 cp = plt.contour(X, Y, Z, colors='white')
    #                 fig.colorbar(cp1)
    #                 plt.clabel(cp, inline=True)
    #                 plt.title("light threshold of %s µmol/m2s" % light)
    #                 ax.set_ylabel('Temperature threshold (°C)')
    #                 ax.set_xlabel('Oxygen threshold (mg/L)')
    #                 plt.savefig("%s/habitat_treshold/contourplot_Areaday_lake_%s_light_%s_time%s.png" % (
    #                 self.output_folder, lake, light, self.time))
    #
    #                 Z = NTGdays["%s" % lake]
    #                 fig, ax = plt.subplots(1, 1, figsize=(10, 8))
    #                 cp1 = ax.contourf(X, Y, Z,vmin=vminNTG["%s" % lake],vmax=vmaxNTG["%s" % lake])
    #                 cp = plt.contour(X, Y, Z, colors='white')
    #                 fig.colorbar(cp1)
    #                 plt.clabel(cp, inline=True, fmt='%.2e')
    #                 plt.title("light threshold of %s µmol/m2s" % light)
    #                 ax.set_ylabel('Temperature threshold (°C)')
    #                 ax.set_xlabel('Oxygen threshold (mg/L)')
    #                 plt.savefig("%s/habitat_treshold/contourplot_NTGdays_lake_%s_light_%s_time%s.png" % (
    #                 self.output_folder, lake, float(light), self.time))
    #                 plt.close('all')
    #
    #
    #     final4 = True
    #     for lake in lakes_data["lake_id"]:
    #         fig, axs = plt.subplots(4, 4, figsize=(15, 15))
    #         for light in range(0, len(light_list)):
    #             if final4:
    #                 if light_list[light] != int(light_list[light]):
    #                     light_list[light] = "%s0" % light_list[light]
    #             Z = AreaDays_o2["%s" % light_list[light]]["%s" % lake]
    #             axis_x = int(np.floor(light / 4))
    #             axis_y = int(light - (4 * axis_x))
    #             print(axs)
    #             ax = axs[axis_x][axis_y]
    #             cp1 = ax.contourf(X, Y, Z,vmin=vminAD["%s" % lake],vmax=vmaxAD["%s" % lake])
    #             # fig.colorbar(cp)  # Add a colorbar to a plot
    #             cp = ax.contour(X, Y, Z, colors='white')
    #             plt.clabel(cp, inline=True, fmt='%.2e')
    #             ax.title.set_text("Light threshold: %s µmol/m2s" % light_list[light])
    #
    #         final4 = False
    #
    #         left = 0.15  # the left side of the subplots of the figure
    #         right = 0.9  # the right side of the subplots of the figure
    #         bottom = 0.05  # the bottom of the subplots of the figure
    #         top = 0.8  # the top of the subplots of the figure
    #         wspace = 0.4 # the amount of width reserved for blank space between subplots
    #         hspace = 0.5  # the amount of height reserved for white space between subplots
    #
    #         #axs[3][3].set_visible(False)
    #         plt.subplots_adjust(right=0.85, left=left, wspace=wspace, hspace=hspace)
    #         cbar_ax = fig.add_axes([0.9, 0.15, 0.05, 0.7])
    #         plt.colorbar(cp1, cax=cbar_ax)
    #         # plt.subplots_adjust(left, bottom, right, top, wspace, hspace)
    #         fig.text(0.5, 0.01, 'Oxygen threshold mg/L', ha='center', va='center')
    #         fig.text(0.01, 0.5, 'Temperature threshold (°C)', ha='center', va='center', rotation='vertical')
    #         plt.savefig(
    #             "%s/habitat_treshold/contourplot_AreaDays_lake_%s_light_time%s.png" % (self.output_folder, lake, self.time))
    #         plt.close('all')
    #
    #         fig, axs = plt.subplots(4, 4, figsize=(15, 15))
    #         for light in range(0, len(light_list)):
    #             Z = Areaday_o2["%s" % light_list[light]]["%s" % lake]
    #             axis_x = int(np.floor(light / 4))
    #             axis_y = int(light - (4 * axis_x))
    #             print(axs)
    #             ax = axs[axis_x][axis_y]
    #             ax.contourf(X, Y, Z,vmin=vminAd["%s" % lake],vmax=vmaxAd["%s" % lake])
    #             # fig.colorbar(cp)  # Add a colorbar to a plot
    #             cp = ax.contour(X, Y, Z, colors='white')
    #             plt.clabel(cp, inline=True, fmt='%.2e')
    #             ax.title.set_text("Light threshold: %s µmol/m2s" % light_list[light])
    #
    #         # axs[3][3].set_visible(False)
    #
    #         plt.subplots_adjust(right=0.9)
    #         cbar_ax = fig.add_axes([0.95, 0.05, 0.05, 0.7])
    #         plt.colorbar(cp1, cax=cbar_ax)
    #         #plt.subplots_adjust(left, bottom, right, top, wspace, hspace)
    #         fig.text(0.5, 0.01, 'Oxygen threshold mg/L', ha='center', va='center')
    #         fig.text(0.01, 0.5, 'Temperature threshold (°C)', ha='center', va='center', rotation='vertical')
    #         plt.savefig(
    #             "%s/habitat_treshold/contourplot_Areaday_lake_%s_light_time%s.png" % (self.output_folder, lake, self.time))
    #         plt.close('all')
    #
    #         fig, axs = plt.subplots(4, 4, figsize=(15, 15))
    #         for light in range(0, len(light_list)):
    #             Z = NTGdays_o2["%s" % light_list[light]]["%s" % lake]
    #             axis_x = int(np.floor(light / 4))
    #             axis_y = int(light - (4 * axis_x))
    #             print(axs)
    #             ax = axs[axis_x][axis_y]
    #             ax.contourf(X, Y, Z,vmin=vminNTG["%s" % lake],vmax=vmaxNTG["%s" % lake])
    #             # fig.colorbar(cp)  # Add a colorbar to a plot
    #             cp = ax.contour(X, Y, Z, colors='white')
    #             plt.clabel(cp, inline=True, fmt='%.2e')
    #             ax.title.set_text("Light threshold: %s µmol/m2s" % light_list[light])
    #
    #         # axs[3][3].set_visible(False)
    #
    #         plt.subplots_adjust(right=0.8,left=left,bottom=bottom,top=top)
    #         cbar_ax = fig.add_axes([0.9, 0.15, 0.05, 0.7])
    #         plt.colorbar(cp1, cax=cbar_ax)
    #         #plt.subplots_adjust(left, bottom, right, top, wspace, hspace)
    #         fig.text(0.5, 0.01, 'Oxygen threshold (mg/L)', ha='center', va='center')
    #         fig.text(0.01, 0.5, 'Temperature threshold (°C)', ha='center', va='center', rotation='vertical')
    #         plt.savefig(
    #             "%s/habitat_treshold/contourplot_NTGdays_lake_%s_light_time%s.png" % (self.output_folder, lake, self.time))
    #         plt.close('all')

    def density_plot(self,summary,calibration,old):
        colorpalette = ["orange", "red", "forestgreen"]

        # sns.set_style("ticks", {"xtick.major.size": 8, "ytick.major.size": 8})
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")

        plt.grid(False)
        sns.distplot(summary["rmseT"], hist=False, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     color=colorpalette[0], label=None)
        plt.ylabel("Density")
        plt.xlabel("RMSE (Celcius)")
        plt.savefig("densityrmseTc%so%s.png" % (calibration, old))
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")

        plt.grid(False)
        sns.distplot(summary["rmseO"], hist=False, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     color=colorpalette[1], label=None)
        plt.ylabel("Density")
        plt.xlabel("RMSE (mg*L-1)")
        plt.savefig("densityrmseOc%so%s.png" % (calibration, old))
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")

        plt.grid(False)
        sns.distplot(summary["rmseS"], hist=False, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     color=colorpalette[2], label=None)
        plt.ylabel("Density")
        plt.xlabel("RMSE (m)")
        plt.savefig("densityrmseSc%so%s.png" % (calibration, old))
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        sns.distplot(summary["nrmseT"], hist=False, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     color=colorpalette[0], label="Temperature")
        sns.distplot(summary["nrmseO"], hist=False, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     color=colorpalette[1], label="Oxygen")
        sns.distplot(summary["nrmseS"], hist=False, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     color=colorpalette[2], label="Secchi depth")
        plt.ylabel("Density")
        plt.xlabel("Normalised RMSE")
        plt.savefig("densityc%so%s.png" % (calibration, old))
        plt.close()

        dataall = pd.read_csv("2017SwedenList.csv", encoding='ISO-8859-1')
        dataall["area"] = dataall["area"]
        dataall["volume"] = dataall["volume"] * 1e-9
        dataall["sedimentArea"] = dataall["sedimentArea"] * 1e-18
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        sns.distplot(dataall["area"], color="black", hist=True, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     label=None, norm_hist=True)
        plt.xscale('log')

        plt.ylabel("Density")
        plt.xlabel("Area ( km2)")
        plt.savefig("area.png")
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        sns.distplot(dataall["depth"], color="black", hist=True, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     label=None)
        plt.ylabel("Density")
        plt.xlabel("Depth (m)")
        plt.savefig("depth.png")
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        sns.distplot(dataall["longitude"], color="black", hist=True, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     label=None)
        plt.ylabel("Density")
        plt.xlabel("Longitude")
        plt.savefig("lon.png")
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        sns.distplot(dataall["latitude"], color="black", hist=True, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     label=None)
        plt.ylabel("Density")
        plt.xlabel("Latitude")
        plt.savefig("latitude.png")
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        sns.distplot(dataall["volume"], color="black", hist=False, kde=True, kde_kws={'shade': True, 'linewidth': 3},
                     label=None)
        plt.ylabel("Density")
        plt.xlabel("Volume (km3)")
        plt.savefig("vol.png")
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        mean = dataall["Mean"].dropna()
        sns.distplot(mean, color="black", hist=True, kde=True, kde_kws={'shade': True, 'linewidth': 3}, label=None)
        plt.ylabel("Density")
        plt.xlabel("Mean depth (m)")
        plt.savefig("mean.png")
        plt.close()
        sns.set(font_scale=2)
        plt.figure(figsize=(10, 8))
        sns.set_style("ticks")
        plt.grid(False)
        sns.distplot(dataall["sedimentArea"], color="black", hist=False, kde=True,
                     kde_kws={'shade': True, 'linewidth': 3}, label=None)
        plt.ylabel("Density")
        plt.xlabel("sedimentArea (x10^12 km\u00b2) ")
        plt.savefig("sediment.png")
        plt.close()

    def observation_plot(self,finalT,T_final):
        fig1 = plt.figure(figsize=(10, 5))
        ax = plt.subplot(111)
        fig1 = sns.scatterplot(data=finalT, x="dates", y="obsT", hue="lakepath", palette="dark")
        box = ax.get_position()
        ax.set_position([box.x0, box.y0 + box.height * 0.1,
                         box.width, box.height * 0.9])

        # Put a legend below current axis
        ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.05),
                  fancybox=True, shadow=True, ncol=5)
        plt.xlim([finalT['dates'].min(), finalT['dates'].max()])
        plt.savefig("T_obs.png")
        finalT = T_final.dropna(subset=['obsO'])
        fig2 = plt.figure(figsize=(10, 5))
        ax = plt.subplot(111)
        fig2 = sns.scatterplot(data=T_final, x="dates", y="obsO", hue="lakepath", palette="dark")
        # Shrink current axis's height by 10% on the bottom
        box = ax.get_position()
        ax.set_position([box.x0, box.y0 + box.height * 0.1,
                         box.width, box.height * 0.9])

        # Put a legend below current axis
        ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.05),
                  fancybox=True, shadow=True, ncol=5)
        plt.xlim([finalT['dates'].min(), finalT['dates'].max()])
        plt.savefig("O2_obs.png")

    def violin_plot(self,rcp="85", lakes_list1="2017SwedenList.csv"):
        import seaborn as sns
        # lake = lakes_list[lake_number]
        lakes = pd.read_csv(lakes_list1, encoding='ISO-8859-1')
        lakes_data = lakes.set_index("lake_id").to_dict()
        lakes_list = list(lakes_data.get("name").keys())
        sns.set_color_codes("colorblind")
        sns.set_context("paper", 2.5)
        variables = [["Tzt.csv", "Change in Surface Temperature ($^\circ$C)"],
                     ["O2zt.csv", "Change in Bottom Oxygen\n Concentration (mg m-2)"],
                     ["His.csv", "Change in Ice Cover Duration (day)"]]
        model_data = [
            ["model", "lake", "volume", "depth", "scenario", variables[0][1], variables[1][1], variables[2][1]]]
        lakesss_data = pd.DataFrame(columns=["lake", "model", "volume", "depth",
                                             "dateM", "dateD", "historicalT", "rcp45T", "rcp85T", "diff45T", "diff85T",
                                             "historicalO", "rcp45O", "rcp85O", "diff45O", "diff85O",
                                             "historicalI", "rcp45I", "rcp85I", "diff45I", "diff85I"])
        # kernel = [["model", "lake", "scenario", variables[0][1], variables[1][1], variables[2][1]]]
        aaaa = 0
        # for modelid in [2]:
        #     m1, m2 = models[modelid]
        #     # if 1==1:
        #     try:
        #
        #         n = 1
        #         if aaaa == 0:
        #             lakess_data = pd.DataFrame(
        #                 columns=["lake", "model", "volume", "depth",
        #                          "dateM", "dateD", "historicalT", "rcp45T", "rcp85T",
        #                          "diff45T", "diff85T",
        #                          "historicalO", "rcp45O", "rcp85O", "diff45O", "diff85O",
        #                          "historicalI", "rcp45I", "rcp85I", "diff45I", "diff85I"])
        #         else:
        #             lakesss_data = lakesss_data.append(lakess_data, ignore_index=True)
        #         aaaa = 0
        #         lake_number = -1
        #         # if 1==1:
        #         for lake in lakes_list:
        #             lake_number += 1
        #             lake_data = pd.DataFrame(
        #                 columns=["lake", "model", "volume", "depth",
        #                          "dateM", "dateD", "historicalT", "rcp45T", "rcp85T",
        #                          "diff45T", "diff85T",
        #                          "historicalO", "rcp45O", "rcp85O", "diff45O", "diff85O",
        #                          "historicalI", "rcp45I", "rcp85I", "diff45I", "diff85I"])
        #             print(m2, lake, n, lake_number)
        #             n += 1
        #
        #             volume = lakes_data.get('volume').get(lake)
        #             depth = lakes_data.get('depth').get(lake)
        #             eh = lakes_data.get('ebhex').get(lake)
        #             eh = eh[2:] if eh[:2] == '0x' else eh
        #             while len(eh) < 6:
        #                 eh = '0' + eh
        #             d1, d2, d3 = eh[:2], eh[:4], eh[:6]
        #
        #             if rcp == "85":
        #                 sce = 8
        #             else:
        #                 sce = 5
        #             for scenarioid in [1, sce]:
        #                 exA, y1A, exB, y1B = scenarios[scenarioid]
        #                 # y2A = y1A + 4
        #                 y2B = y1B + 4
        #                 outdir = os.path.join(output_path, d1, d2, d3,
        #                                       'EUR-11_%s_%s-%s_%s_%s0101-%s1231' % (m1, exA, exB, m2, y1A, y2B))
        #
        #                 # lakeinfo = LakeInfo(lakes_list[lake_number], list(lakes["lake_id"])[lake_number],
        #                 #                     list(lakes["subid"])[lake_number], list(lakes["ebhex"])[lake_number],
        #                 #                     list(lakes["area"])[lake_number], list(lakes["depth"])[lake_number],
        #                 #                     list(lakes["Mean"])[lake_number],
        #                 #                     list(lakes["longitude"])[lake_number], list(lakes["latitude"])[lake_number],
        #                 #                     list(lakes["volume"])[lake_number], scenarioid, scenarioid=scenarioid,
        #                 #                     modelid=modelid)
        #
        #                 # lake.variables_by_depth()
        #                 # lakeinfo.runlake(modelid,scenarioid)
        #
        #                 if scenarioid == 1:
        #                     dstart = date(y1A, 1, 1)
        #                     dend = date(y2B, 12, 31)
        #
        #                     # this will give you a list containing all of the dates
        #                     dd = [dstart + timedelta(days=x) for x in range((dend - dstart).days + 1)]
        #
        #                     date_stringsM = [d.strftime('%m') for d in dd]
        #                     date_stringsD = [d.strftime('%d') for d in dd]
        #                     lake_data["dateM"] = date_stringsM
        #                     lake_data["dateD"] = date_stringsD
        #                     lake_data["lake"] = lake
        #                     lake_data["model"] = modelid
        #                     lake_data["volume"] = volume
        #                     lake_data["depth"] = depth
        #
        #                 for variable in [0, 1, 2]:
        #                     data = pd.read_csv(os.path.join(outdir, variables[variable][0]), header=None)
        #
        #                     if variable == 0:
        #                         lake_data["%sT" % exA] = data[0]
        #                     elif variable == 1:
        #                         lake_data["%sO" % exA] = data.iloc[:, -1] * 0.001
        #                     else:
        #                         icecoverduration = (data[6].sum()) / 10
        #                         lake_data["%sI" % exA] = icecoverduration
        #
        #             data_summary = lake_data.mean()
        #             for rcp in [rcp]:
        #                 for letter in ["T", "O", "I"]:
        #                     data_summary["diff%s%s" % (rcp, letter)] = data_summary["rcp%s%s" % (rcp, letter)] - \
        #                                                                data_summary[
        #                                                                    "historical%s" % letter]
        #             if aaaa == 0:
        #                 lakess_data = lake_data
        #                 aaaa += 1
        #             else:
        #                 lakess_data = lakess_data.append(lake_data, ignore_index=True)
        #
        #             for rcp in [rcp]:
        #                 model_code = {1: 'KNM',
        #                               2: 'DMI',
        #                               3: 'MPI',
        #                               4: 'MOH',
        #                               5: 'IPS',
        #                               6: 'CNR'}
        #                 model_data.append(
        #                     [model_code.get(modelid), lake, volume, depth, "rcp%s" % rcp, data_summary["diff%sT" % rcp],
        #                      data_summary["diff%sO" % rcp], data_summary["diff%sI" % rcp]])
        #
        #     #
        #     except:
        #         print("model %s doesnt exist" % (m1 + m2))
        #
        import seaborn as sns
        from matplotlib.cm import ScalarMappable
        import matplotlib.pyplot as plt
        #
        # headers = model_data.pop(0)
        # timestr = time.strftime("%Y%m%d-%H%M%S")
        # final_data = pd.DataFrame(model_data, columns=headers)
        # lakesss_data.to_csv("annually_average_T_Ice_cover_Oxygen_originall_%s.csv" % rcp)
        # final_data.to_csv("annually_average_T_Ice_cover_Oxygen_%s.csv" % rcp)
        # cpal = sns.color_palette('coolwarm_r', n_colors=3, desat=1.)
        final_data = pd.read_csv("annually_average_T_Ice_cover_Oxygen_%s.csv" % rcp)
        # final_data['Lake_group'] = 2
        # final_data.loc[final_data['volume'] < 1.1e8, 'Lake_group'] = 1
        # final_data.loc[final_data['volume'] > 1.1e10, 'Lake_group'] = 3
        plotT = sns.catplot(x="model", y=variables[0][1], col="scenario", data=final_data, kind="violin", color="blue",
                            split=True,
                            height=8, aspect=.8)
        # sns.set(style="ticks", palette="colorblind")
        # g = sns.FacetGrid(final_data, col="model", sharey=False, size=15, aspect=.5)
        # g = g.map(sns.violinplot, "model", "Change in Surface Temperature ($^\circ$C)", "scenario", linewidth=1,palette={"rcp45": "lightblue", "rcp85": "darkblue"},
        #           scale="area", split=True, width=0.75).despine(left=True)
        # sns.stripplot(x="model", y = "Change in Surface Temperature ($^\circ$C)", data=final_data[final_data['scenario']=="rcp45"],alpha=0.8, jitter=.1, edgecolor='black',  size=8, hue='Lake_group',
        #               palette=cpal,linewidth=1,marker='s')
        # for ax in g.axes.flatten():
        #     ax.collections[0].set_edgecolor('k')
        #     ax.collections[1].set_edgecolor('k')
        # #sns.plt.show()
        # plt.savefig(os.path.join(output_path, "violinT1.png" ))
        # plt.close()
        # sns.set(style="ticks", palette="colorblind")
        # g = sns.FacetGrid(final_data, col="model", sharey=False, size=15, aspect=.5)
        # g = g.map(sns.violinplot, "model", "Change in Ice Cover Duration (day)", "scenario", linewidth=1,
        #           palette={"rcp45": "lightgreen", "rcp85": "darkgreen"},
        #           scale="area", split=True, width=0.75).despine(left=True)
        # sns.stripplot(x="model", y="Change in Ice Cover Duration (day)",
        #               data=final_data[final_data['scenario'] == "rcp45"], alpha=0.8, jitter=.1, edgecolor='black', size=8,
        #               hue='Lake_group',
        #               palette=cpal, linewidth=1, marker='s')
        # for ax in g.axes.flatten():
        #     ax.collections[0].set_edgecolor('k')
        #     ax.collections[1].set_edgecolor('k')
        # # sns.plt.show()
        # plt.savefig(os.path.join(output_path, "violinS1.png"))
        # sns.set(style="ticks", palette="colorblind")
        # g = sns.FacetGrid(final_data, col="model", sharey=False, size=15, aspect=.5)
        # g = g.map(sns.violinplot, "model", "Change in Bottom Oxygen Concentration (mg m-2)", "scenario", linewidth=1,
        #           palette={"rcp45": "mistyrose", "rcp85": "darkred"},
        #           scale="area", split=True, width=0.75).despine(left=True)
        # sns.stripplot(x="model", y="Change in Bottom Oxygen Concentration (mg m-2)",
        #               data=final_data[final_data['scenario'] == "rcp45"], alpha=0.8, jitter=.1, edgecolor='black', size=8,
        #               hue='Lake_group',
        #               palette=cpal, linewidth=1, marker='s')
        # for ax in g.axes.flatten():
        #     ax.collections[0].set_edgecolor('k')
        #     ax.collections[1].set_edgecolor('k')
        # # sns.plt.show()
        # plt.savefig(os.path.join(output_path, "violinO1.png"))
        # plt.close()
        plotT.savefig(os.path.join(self.output_path, "violinT1_%s_%s.png" % (rcp, self.timestr)))
        plt.close(fig=plotT)
        # plotI2 = sns.violinplot(x="model", y=variables[0][1], col="scenario", data=final_data, inner=None, color=".8")
        # ploT2 = sns.stripplot(x="model", y=variables[0][1], data=final_data, hue="volume", palette="viridis")
        # # Colormap for comparison
        # # Data
        # y_min = min(final_data["volume"])
        # y_max = max(final_data["volume"])
        # cmap = plt.get_cmap("viridis")
        # norm = plt.Normalize(y_min, y_max)
        # sm = ScalarMappable(norm=norm, cmap=cmap)
        # sm.set_array([])
        # cbar = fig.colorbar(sm, ax=ax2)
        # plotT2.savefig(os.path.join(self.output_path, "violinT1_%s_%s.png" % (rcp, self.timestr)))

        print("save T")
        plotO = sns.catplot(x="model", y=variables[1][1], col="scenario", data=final_data, kind="violin",
                            color="red",
                            split=True,
                            height=8, aspect=.9)
        plotO.savefig(os.path.join(self.output_path, "violinO1_%s_%s.png" % (rcp, self.timestr)))
        print("save O")
        plotI = sns.catplot(x="model", y=variables[2][1], col="scenario", data=final_data, kind="violin",
                            color="forestgreen", split=True,
                            height=8, aspect=.7)
        plotI.savefig(os.path.join(self.output_path, "violinI1_%s_%s.png" % (rcp, self.timestr)))
        print("save Ice")
        print("end violin")

    def violin_plot45(self,lakes_list1="2017SwedenList.csv", model_data=pd.DataFrame()):
        # lake = lakes_list[lake_number]
        lakes = pd.read_csv(lakes_list1, encoding='ISO-8859-1')
        lakes_data = lakes.set_index("lake_id").to_dict()
        lakes_list = list(lakes_data.get("name").keys())
        sns.set_color_codes("colorblind")
        sns.set_context("paper", 2.5)
        variables = [["Tzt.csv", "Change in Surface Temperature ($^\circ$C)"],
                     ["O2zt.csv", "Change in Bottom Oxygen\n Concentration (mg m-2)"],
                     ["His.csv", "Change in Ice Cover Duration (day)"]]
        # model_data = [["model", "lake", "volume", "depth", "scenario", variables[0][1], variables[1][1], variables[2][1]]]
        # lakesss_data = pd.DataFrame(columns=["lake", "model", "volume", "depth",
        #                                      "dateM", "dateD", "historicalT", "rcp45T", "rcp85T", "diff45T", "diff85T",
        #                                      "historicalO", "rcp45O", "rcp85O", "diff45O", "diff85O",
        #                                      "historicalI", "rcp45I", "rcp85I", "diff45I", "diff85I"])
        # # kernel = [["model", "lake", "scenario", variables[0][1], variables[1][1], variables[2][1]]]
        # aaaa = 0
        # for modelid in [2]:
        #     m1, m2 = models[modelid]
        #     # if 1==1:
        #     if 1==1:##try:
        #
        #         n = 1
        #         if aaaa == 0:
        #             lakess_data = pd.DataFrame(
        #                 columns=["lake", "model", "volume", "depth",
        #                          "dateM", "dateD", "historicalT", "rcp45T", "rcp85T",
        #                          "diff45T", "diff85T",
        #                          "historicalO", "rcp45O", "rcp85O", "diff45O", "diff85O",
        #                          "historicalI", "rcp45I", "rcp85I", "diff45I", "diff85I"])
        #         else:
        #             lakesss_data = lakesss_data.append(lakess_data, ignore_index=True)
        #         aaaa = 0
        #         lake_number = -1
        #         # if 1==1:
        #         for lake in lakes_list:
        #             lake_number += 1
        #             lake_data = pd.DataFrame(
        #                 columns=["lake", "model", "volume", "depth",
        #                          "dateM", "dateD", "historicalT", "rcp45T", "rcp85T",
        #                          "diff45T", "diff85T",
        #                          "historicalO", "rcp45O", "rcp85O", "diff45O", "diff85O",
        #                          "historicalI", "rcp45I", "rcp85I", "diff45I", "diff85I"])
        #             print(m2, lake, n, lake_number)
        #             n += 1
        #
        #             volume = lakes_data.get('volume').get(lake)
        #             depth = lakes_data.get('depth').get(lake)
        #             eh = lakes_data.get('ebhex').get(lake)
        #             eh = eh[2:] if eh[:2] == '0x' else eh
        #             while len(eh) < 6:
        #                 eh = '0' + eh
        #             d1, d2, d3 = eh[:2], eh[:4], eh[:6]
        #
        #             for scenarioid in [1, 5, 8]:
        #                 exA, y1A, exB, y1B = scenarios[scenarioid]
        #                 # y2A = y1A + 4
        #                 y2B = y1B + 4
        #                 outdir = os.path.join(output_path, d1, d2, d3,
        #                                       'EUR-11_%s_%s-%s_%s_%s0101-%s1231' % (m1, exA, exB, m2, y1A, y2B))
        #
        #                 lakeinfo = LakeInfo(lakes_list[lake_number], list(lakes["lake_id"])[lake_number],
        #                                     list(lakes["subid"])[lake_number], list(lakes["ebhex"])[lake_number],
        #                                     list(lakes["area"])[lake_number], list(lakes["depth"])[lake_number],
        #                                     list(lakes["Mean"])[lake_number],
        #                                     list(lakes["longitude"])[lake_number], list(lakes["latitude"])[lake_number],
        #                                     list(lakes["volume"])[lake_number], scenarioid, scenarioid=scenarioid,
        #                                     modelid=modelid)
        #
        #                 # lake.variables_by_depth()
        #                 # lakeinfo.runlake(modelid,scenarioid)
        #
        #                 if scenarioid == 1:
        #                     dstart = date(y1A, 1, 1)
        #                     dend = date(y2B, 12, 31)
        #
        #                     # this will give you a list containing all of the dates
        #                     dd = [dstart + timedelta(days=x) for x in range((dend - dstart).days + 1)]
        #
        #                     date_stringsM = [d.strftime('%m') for d in dd]
        #                     date_stringsD = [d.strftime('%d') for d in dd]
        #                     lake_data["dateM"] = date_stringsM
        #                     lake_data["dateD"] = date_stringsD
        #                     lake_data["lake"] = lake
        #                     lake_data["model"] = modelid
        #                     lake_data["volume"] = volume
        #                     lake_data["depth"] = depth
        #
        #                 for variable in [0, 1, 2]:
        #                     data = pd.read_csv(os.path.join(outdir, variables[variable][0]), header=None)
        #
        #                     if variable == 0:
        #                         lake_data["%sT" % exA] = data[0]
        #                     elif variable == 1:
        #                         lake_data["%sO" % exA] = data.iloc[:, -1] * 0.001
        #                     else:
        #                         icecoverduration = (data[6].sum()) / 10
        #                         lake_data["%sI" % exA] = icecoverduration
        #
        #             data_summary = lake_data.mean()
        #             lake_data.to_csv("annually_average_T_Ice_cover_Oxygen_test.csv")
        #             for rcp in ["45", "85"]:
        #                 for letter in ["T", "O", "I"]:
        #                     data_summary["diff%s%s" % (rcp, letter)] = data_summary["rcp%s%s" % (rcp, letter)] - \
        #                                                                data_summary[
        #                                                                    "historical%s" % letter]
        #                     if letter == "I":
        #                         data_summary["ice%s%s" % (rcp, letter)] = data_summary["rcp%s%s" % (rcp, letter)]
        #                         data_summary["icehisto%s" % (letter)] = data_summary["historical%s" % letter]
        #             if aaaa == 0:
        #                 lakess_data = lake_data
        #                 aaaa += 1
        #             else:
        #                 lakess_data = lakess_data.append(lake_data, ignore_index=True)
        #
        #             for rcp in ["45", "85"]:
        #                 model_code = {1: 'KNM',
        #                               2: 'DMI',
        #                               3: 'MPI',
        #                               4: 'MOH',
        #                               5: 'IPS',
        #                               6: 'CNR'}
        #                 model_data.append(
        #                     [model_code.get(modelid), lake, volume, depth, "rcp%s" % rcp, data_summary["diff%sT" % rcp],
        #                      data_summary["diff%sO" % rcp], data_summary["diff%sI" % rcp]])
        #
        #     #
        #     # except:
        #     #     print("model %s doesnt exist" % (m1 + m2))
        headers = model_data.pop(0)
        # timestr = time.strftime("%Y%m%d-%H%M%S")
        final_data = pd.DataFrame(model_data, columns=headers)
        # lakesss_data.to_csv("annually_average_T_Ice_cover_Oxygen_originall.csv")
        # final_data.to_csv("annually_average_T_Ice_cover_Oxygen.csv")
        plotT = sns.catplot(x="model", y=variables[0][1], col="scenario", data=final_data, kind="violin",
                            color="orange",
                            split=True,
                            height=8, aspect=.8)
        plotT.savefig(os.path.join(self.output_path, "violinT4589_%s.png" % (self.timestr)))
        print("save T")
        plotO = sns.catplot(x="model", y=variables[1][1], col="scenario", data=final_data, kind="violin", color="red",
                            split=True,
                            height=8, aspect=.9)
        plotO.savefig(os.path.join(self.output_path, "violinO4589_%s.png" % (self.timestr)))
        print("save O")
        plotI = sns.catplot(x="model", y=variables[2][1], col="scenario", data=final_data, kind="violin",
                            color="forestgreen", split=True,
                            height=8, aspect=.7)
        plotI.savefig(os.path.join(self.output_path, "violinI4589_%s.png" % (self.timestr)))
        print("save Ice")
        print("end violin")

    def density_plot2(self,final_data,variables,rcpscenarios=['rcp45', 'rcp85']):

        # element in fictionnary = [0 = {color for each scenario}, 1 = "column name for variable/x-axis name"]
        variables_info = {"oxygen":[{'rcp45': '#C9C9C9', 'rcp85':'#C00000'},'Change in Bottom Oxygen Concentration'],
                          "temperature":[{'rcp45': '#C9C9C9', 'rcp85':'#2A4D60'},'Change in Surface Temperature'],
                          "ice":[{'rcp45': '#C9C9C9', 'rcp85':'#70AD47'},'Change in Ice Cover Duration']}
        number_row = len(variables)

        fig, axes = plt.subplots(number_row, 1, gridspec_kw={
                           'width_ratios': [2],
                           'height_ratios': [1]*number_row})
        # order given to variables = order subplots
        for variable in range(0,len(variables)):
            for scenario in rcpscenarios:
                #Subset the specific scenario
                subset = final_data[final_data['scenario'] == scenario]
                color = variables_info[variables[variable]][0][scenario]
                labelscenario = "RCP 4.5"
                if scenario == "rcp45":
                    labelscenario = "RCP 4.5"
                else:
                    labelscenario = "RCP 8.5"
                # Draw the density plot
                sns.distplot(subset[variables_info[variables[variable]][1]], hist=False, kde=True,
                             kde_kws={'linewidth': 3}, color=color, label=labelscenario, ax=axes[variable])
                axes[variable].set_ylabel("Density")
                if variable == 0:
                    axes[variable].set_xlabel("Change in surface water temp. (°C)")
                elif variable == 1:
                    axes[variable].set_xlabel("Change in bottom DO concentration\n (mg $\mathregular{L^{-1}}$)")
                else:
                    axes[variable].set_xlabel("Change in ice cover duration\n (days per decade)")

        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.MEDIUM_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title

        # fig.subplots_adjust(hspace=0.3)
        fig.set_size_inches(self.width, self.height)
        plt.tight_layout()
        fig.subplots_adjust(left=0.25, right=.95, hspace=0.4)
        fig.set_size_inches(self.width, self.height)

        plt.rc('font', size=self.SMALL_SIZE)  # controls default text sizes
        plt.rc('axes', titlesize=self.MEDIUM_SIZE)  # fontsize of the axes title
        plt.rc('axes', labelsize=self.BIGGER_SIZE)  # fontsize of the x and y labels
        plt.rc('xtick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('ytick', labelsize=self.SMALL_SIZE)  # fontsize of the tick labels
        plt.rc('legend', fontsize=self.SMALL_SIZE)  # legend fontsize
        plt.rc('figure', titlesize=self.MEDIUM_SIZE)  # fontsize of the figure title
        plt.rcParams["font.family"] = self.font_family

        final_data.to_csv("final_density.csv")
        print(os.path.join(self.output_folder,"figure8_%s%s.svg"%(self.time,self.font_family)))
        plt.savefig(os.path.join(self.output_folder,"figure8_%s%s.svg"%(self.time,self.font_family)))
        plt.savefig(os.path.join(self.output_folder, "figure8_%s%s.eps" % (self.time, self.font_family)),format='eps')
        plt.savefig(os.path.join(self.output_folder,"figure8_%s%s.png"%(self.time,self.font_family)))
        plt.close()
        plt.clf()


def violin_parallel():
    """
    Simple function to call a parallel calibration of all lakes.
    :return: None
    """
    print(num_cores)
    # for lake in lakes_list:
    #     run_calibrations(lake)
    Parallel(n_jobs=num_cores, verbose=10)(delayed(violin_plot)(lakenumber) for lakenumber in range(0, len(lakes_list)))




def exploratory(lakelist="2017SwedenList_only_validation_12lakes.csv"):
    df = pd.read_csv(lakelist, encoding='ISO-8859-1')
    df['Lake_group'] = 2
    df.loc[df['volume'] < 1.0e7, 'Lake_group'] = 1
    df.loc[df['volume'] > 5.0e9, 'Lake_group'] = 3
    df[['area','depth','longitude','latitude','Mean','volume','Turnover',
        'sedimentArea','Subcatchment_Area','Catchment_Area','C.L','SC.L','Lake_group']].hist(bins=20,figsize=(15, 15))
    #plt.show()
    plt.savefig("explo.png")

    continous_var = ['area','depth','longitude','latitude','Mean','volume','Turnover',
        'sedimentArea','Subcatchment_Area','Catchment_Area','C.L','SC.L']
    continous_var_label = ['Area', 'Max_Depth', 'Longitude', 'Latitude', 'Mean_depth', 'Volume', 'Residence_Time',
                           'Sediment_Area', 'Subcatchment_Area', 'Catchment_Area', 'Catchment_Lake_ratio',
                           'Subcatchment_Lake_ratio']
    df2 = df[continous_var].describe()
    df2.to_csv("describe.csv")

    df['Lake_group'] = 2
    df.loc[df['volume'] < 1.0e7, 'Lake_group'] = 1
    df.loc[df['volume'] > 5.0e9, 'Lake_group'] = 3
   # sns.scatterplot(x="area", y="depth", data=df, hue="Lake_group")
    #plt.savefig("explo2.png")
    sns.pairplot(df[['area','depth','longitude','latitude','Mean','volume','Turnover',
        'sedimentArea','Subcatchment_Area','Catchment_Area','C.L','SC.L','Lake_group']], hue="Lake_group")

    plt.savefig("explo2.png")

    plt.figure(figsize=(32, 50))

    for i, col in enumerate(continous_var):
        plt.subplot(12, 4, i * 2 + 1)
        #plt.subplots_adjust(hspace=.25, wspace=.3)

        plt.grid(True)
        plt.title(col)
        sns.kdeplot(df.loc[df["Lake_group"] == 1, col], label="Small", color="green", shade=True, kernel='gau', cut=0)
        sns.kdeplot(df.loc[df["Lake_group"] == 2, col], label="Medium", color="red", shade=True, kernel='gau', cut=0)
        sns.kdeplot(df.loc[df["Lake_group"] == 3, col], label="Large", color="blue", shade=True, kernel='gau', cut=0)
        plt.subplot(12, 4, i * 2 + 2)
        sns.boxplot(y=col, data=df, x="Lake_group", palette=["green", "red","blue"])
    plt.savefig("explo3.png")

    plt.figure(figsize=(32, 50))

    continous_var = ['area', 'depth','longitude','latitude', 'Mean', 'volume', 'Turnover',
                     'sedimentArea', 'Subcatchment_Area', 'Catchment_Area', 'C.L', 'SC.L']
    continous_var_label = ['Area', 'Max_Depth', 'Longitude', 'Latitude', 'Mean_depth', 'Volume', 'Residence_Time',
                     'Sediment_Area', 'Subcatchment_Area', 'Catchment_Area', 'Catchment_Lake_ratio', 'Subcatchment_Lake_ratio']
    palette = iter(sns.husl_palette(len(continous_var)))
    for i, col in enumerate(continous_var):
        plt.subplot(21, 2, i  + 1)
        # plt.subplots_adjust(hspace=.25, wspace=.3)

        plt.grid(True)
        plt.title(col)
        data = df[col]
        #sns.kdeplot(data, color=palette[i], shade=True, kernel='gau', cut=0)
        sns.distplot(data, hist=True, kde=True,
             bins=int(180/5), color = palette[i],
             hist_kws={'edgecolor':'black'},
             kde_kws={'linewidth': 4})
        plt.set_title(continous_var_label[i])

    plt.savefig("explo4.png")

if __name__ == "__main__":
    print("hello")
    SMALL_SIZE = 12
    MEDIUM_SIZE = 14
    BIGGER_SIZE = 16
    sns.set(rc={'figure.figsize': (4.5,5)})
    plt.style.context('seaborn-paper')
    sns.set_style("ticks")
    plt.rc('font', size=SMALL_SIZE)  # controls default text sizes
    plt.rc('axes', titlesize=MEDIUM_SIZE)  # fontsize of the axes title
    plt.rc('axes', labelsize=BIGGER_SIZE)  # fontsize of the x and y labels
    plt.rc('xtick', labelsize=SMALL_SIZE)  # fontsize of the tick labels
    plt.rc('ytick', labelsize=SMALL_SIZE)  # fontsize of the tick labels
    plt.rc('legend', fontsize=MEDIUM_SIZE)  # legend fontsize
    plt.rc('figure', titlesize=MEDIUM_SIZE)  # fontsize of the figure title
    plt.rc("image",cmap='viridis')
    lakes_list1 = "2017SwedenList209.csv"
    # mainfn.comparison_plot_v2(lakes_listcsv=lakes_list1, modelid=2, scenarioid=2, outputfolder=r'F:\output')
    # # mainfn.ice_cover_comparison(lake_list="2017SwedenList.csv")
    # mainfn.violin_plot45(lakes_list1="2017SwedenList.csv", output_path=r"F:\output")
    #
    # mainfn.summary_characteristics_lake(lakes_listcsv=lakes_list1, calibration=True,
    #                              withfig=True, old=False, outputfolder=r'F:\output', new=False)
    temp_list = [1,2, 4, 6, 8, 10, 12, 14, 15]
    # final_data = p
    # d.read_csv("annually_average_T_Ice_cover_Oxygen.csv")
    # Graphics("F:\output_final",height=3*3,font_family="Times New Roman",size=13).density_plot2(final_data,["temperature","oxygen","ice"])
    # Graphics("F:\output_final", height=3 * 3, font_family="Arial",size=12).density_plot2(final_data,["temperature","oxygen","ice"])
    #
    # final_data = pd.read_csv("annually_average_T_Ice_cover_Oxygen.csv")
    # Graphics(r"C:\Users\macot620\Documents").violin_plot45(model_data=model_data)
    # Graphics("F:\output_final", height=3 * 3, font_family="Times New Roman", size=13).density_plot2(final_data,["temperature", "oxygen", "ice"])

    # Graphics("F:\output",width=6.5,height=3*3,font_family="Times New Roman",size=13).timeseries_habitat_by_model()
    # Graphics("F:\output", width=6.5, height=6.5, font_family="Arial", size=12).timeseries_habitat_by_model(listmodels=[2],uninhabitable=False)
    #
    T_list = [0, 3, 6, 9, 12, 15, 18, 21, 24, 25]
    light_list = [0,0.1,0.2,0.5,5,10, 15, 25, 35, 50, 65, 75, 85, 100]

    oxy_list = [0, 3, 3.5, 4, 4.5, 5, 5.5, 6, 6.5, 7, 7.5, 8, 8.5, 9, 9.5, 10]
    label_axis_in_order = {'xlabel':'Light threshold (µmol/(m2s))', 'ylabel':'Temperature threshold (°C)','zlabel':''}
    variables_list_in_order = ['Light','Temperature','Oxygen']
    Graphics(r"F:\output", height=6.5,width=6.5,font_family="Arial",size=9.5).contourplot_temp_vs_light_oxy(light_list,T_list,oxy_list,variables_list_in_order,label_axis_in_order,lakes_list=lakes_list1)
    # #

    #
    # oxygen_list = [3,   3.5,    4,  4.5,    5,  5.5,     6,  6.5,    7, 7.5,   8,   8.5,    9,  9.5,    10, 10.5]
    # #Graphics(r"F:\output").contourplot_temp_vs_light_oxy( temp_list, light_list, oxygen_list)
#     #
#     # light_list = [0.5, 1, 2, 4, 8, 12, 20, 25, 30, 40, 50, 60, 70, 80, 90, 100]
#     # Graphics(r"F:\output").contourplot_temp_vs_light_oxy(oxygen_list,temp_list,light_list,
#     #                                                      ['Oxygen','Temperature','Light'],
#     #                                                      {'xlabel':'Oxygen threshold (mg/L)','ylabel':'Temperature threshold (°C)','zlabel':'Light threshold (µmol/m2s)'},individual=True)
#     #
#
#     temp_list = [*range(0, 26, 1)]
#     temp_list = [0,3,6,9,12,15,18,21,24,25]
#     oxygen_list = [0,3,3.5,4,4.5,5,5.5,6,6.5,7,7.5,8,8.5,9,9.5,10]
#     light_list = [0.5,1,2,4, 8, 12, 16, 20, 25,30,40,50, 60, 70, 80, 90, 100]
#     light_list = [0.5, 1, 2, 4, 8, 12, 16, 20, 25, 30, 40, 50, 60, 70, 80, 90, 100]
#     oxy_list = [0, 3, 3.5, 4, 4.5, 5, 5.5, 6, 6.5, 7, 7.5, 8, 8.5, 9, 9.5, 10]
#
    T_list = [0, 3, 6, 9, 12, 15, 18, 21, 24, 25]
    light_list = [0.5, 15, 25, 35, 50, 65, 75, 85, 100]

#     oxy_list = [0, 3, 3.5, 4, 4.5, 5, 5.5, 6, 6.5, 7, 7.5, 8, 8.5, 9, 9.5, 10]
#     Graphics(r"F:\output").contourplot_temp_vs_light_oxy(light_list, temp_list, oxygen_list,
#                                                          ['Light', 'Temperature', 'Oxygen'],
#                                                          {'xlabel': 'Light threshold (µmol/m2s)',
#                                                           'ylabel': 'Temperature threshold (°C)',
#                                                           'zlabel': 'Oxygen threshold (mg/L)'},individual=False)
# #
#
#     Graphics(r"F:\output").contourplot_temp_vs_light_oxy()
#     Graphics(r"F:\output").contourplot_temp_vs_oxy_light(temp_list, light_list, oxygen_list)
# #
#     label_axis_in_order = {'xlabel':'Light threshold (µmol/(m2s))', 'ylabel':'Temperature threshold (°C)','zlabel':''}
#     variables_list_in_order = {'x':'Light','y':'Temperature','z':'Oxygen'}
#     Graphics(r"F:\output_final").contourplot_temp_vs_light_oxy(light_list,T_list,oxy_list,variables_list_in_order,label_axis_in_order)
#
